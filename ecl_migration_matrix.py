"""
ECL Stage Migration Matrix & Commentary Narrator
--------------------------------------------------
AICA Level 2 Capstone Project
Author : Hitesh Bhadada (Financial Controller)

PURPOSE
-------
Goes beyond an aggregated ECL summary to show LOAN-LEVEL stage migration:
"this loan was Stage 1 last period, Stage 3 this period" - the kind of
bucket movement table a Board/Audit Committee actually wants to interrogate,
with click-through to the underlying customer list.

WHAT THIS SCRIPT DOES (and does NOT do)
----------------------------------------
  1. Reads loan-level data: Period, LoanID, CustomerName, Stage,
     Outstanding_Cr, Provision_Cr  -->  plain arithmetic only.
  2. Builds a "From Stage" (rows) x "To Stage" (columns) MIGRATION MATRIX
     for every pair of consecutive periods - deterministic, auditable.
  3. Writes an Excel workbook where every number in the matrix is a
     CLICKABLE HYPERLINK jumping to a detail sheet listing the exact
     loans/customers behind that number (loan-level drill-down).
  4. Optionally (--ai flag, needs internet + API key) drafts a Board-ready
     narrative paragraph from the ALREADY-COMPUTED aggregate figures only.

This script never computes ECL provisioning itself (that stays governed by
Ind AS 109 / RBI norms in your source systems) and AI is used only to
draft prose from numbers already computed - never to move a loan between
stages or to compute a provision.

HOW IT WORKS
-----------
Keep one file per period end in a data folder. The tool scans that folder,
lists every period it finds, and asks which two you want to compare. Only
the selected periods are analysed and only their sheets are produced -
nothing about the periods is hard-coded.

    <your folder>/
      data/                      <- one file per period end (you maintain this)
        Q1FY25.csv
        Q2FY25.xlsx
        ...
        Q4FY26.csv
      input_archive/             <- created: dated copy of the files used
      output/                    <- created: one timestamped folder per run
        Q4FY25-vs-Q3FY26_2026-08-29_012240/
          ..._ECL_analysis_....xlsx
          ..._commentary_....txt
          run_log_....txt

Files may be CSV, XLSX, XLSM, XLS or XLSB, freely mixed. The period label is
read from the Period column; if that column is absent the file name is used,
so a bare "Q3FY26.csv" works. Period labels are ordered chronologically
(Q1FY25 ... Q4FY26), not alphabetically, so prior-year and cross-year
comparisons behave correctly.

USAGE
-----
    # interactive - scans ./data, lists periods, asks which two to compare
    python ecl_migration_matrix.py

    # a different data folder
    python ecl_migration_matrix.py --data-dir /path/to/period_ends

    # just show what periods are available, then exit
    python ecl_migration_matrix.py --list-periods

    # skip the prompt (scheduled / batch runs)
    python ecl_migration_matrix.py --compare Q4FY25:Q3FY26
    python ecl_migration_matrix.py --compare Q1FY26:Q3FY26 Q2FY26:Q4FY26

    # legacy: one file holding several periods
    python ecl_migration_matrix.py --input-file all_quarters.csv

    # AI-polished narrative (needs internet + ANTHROPIC_API_KEY)
    python ecl_migration_matrix.py --compare Q4FY25:Q3FY26 --ai

REGULATORY SCOPE: NBFC - MIDDLE LAYER (NBFC-ML) ONLY
------------------------------------------------------
All IRACP thresholds and provisioning rates implemented here are those
applicable to an NBFC classified in the Middle Layer under the RBI (NBFC -
IRACP) Directions, 2025. Base Layer and Upper Layer rates are deliberately
NOT implemented. If the entity is re-classified into another layer, the
constants in this script must be revisited before the output is relied upon.

INPUT FILE FORMAT (CSV or Excel: .xlsx, .xlsm, .xls, .xlsb)
--------------------------------------------------------------
    Period          e.g. Q1FY26, Q2FY26, Q3FY26 (chronological order in file)
    LoanID          unique loan identifier
    CustomerName    borrower name (mask/anonymise before use if sharing the
                     file outside your organisation - see DPDPA note below)
    Stage           "Stage 1" / "Stage 2" / "Stage 3"
    Outstanding_Cr  outstanding amount for that loan in that period (Rs. Cr)
    Provision_Cr    ECL provision held for that loan in that period (Rs. Cr)

    For the RBI IRACP disclosure, supply AT MINIMUM:
    DPD             days past due as at that period-end (integer)
                    -> this alone is enough. The tool auto-classifies every
                       loan (Standard / Sub-standard / Doubtful / Loss) and
                       auto-computes its IRACP provision. No manually
                       pre-computed provision figure is ever required.

    FULLY OPTIONAL refinements (the tool runs without either):
    Secured_Value_Cr realisable value of security held. If this column is
                     absent, blank or non-numeric for a loan, that loan is
                     treated as FULLY UNSECURED for doubtful-asset
                     provisioning. Para 32(2)(i) requires 100% provision on
                     the unsecured portion, so assuming nil security can only
                     OVERSTATE the prudential floor, never understate it -
                     the prudent default. Every affected loan is listed on
                     the auto-generated "IRACP_Assumptions" sheet.
    Loss_Flag        "Y" where the asset has been identified as a loss asset
                     by the NBFC, its internal/external auditor, or the
                     Reserve Bank on inspection [para 11(1)]. If absent, no
                     asset is treated as a loss asset - which is correct,
                     since loss status requires positive identification and
                     is NOT derivable from DPD.

OUTPUT
------
    <input>_migration_analysis.xlsx containing:
        - "Summary"                    : period-wise stage totals + commentary
        - "Mig_<P1>_<P2>"              : migration matrix per period-pair,
                                          numbers are hyperlinks
        - "Det_<From>_<To>_<P1>_<P2>"  : loan-level detail behind each cell
        - "IndAS107_35H_<P1>_<P2>"     : STATUTORY DISCLOSURE - loss allowance
                                          reconciliation (opening to closing)
                                          by ECL category, as required by
                                          Ind AS 107 para 35H
        - "RBI_IRACP_Comparison_<P>"   : STATUTORY DISCLOSURE - comparison of
                                          IRACP provisioning vs Ind AS 109 ECL
                                          and Impairment Reserve requirement,
                                          per RBI's NBFC IRACP Directions, 2025
                                          (generated whenever a DPD column exists)
        - "IRACP_Assumptions_<P>"      : where security data was missing and
                                          what was assumed in its place
        - "IRACP_LoanWorking_<P>"      : per-loan IRACP audit trail

SOFTWARE DEPENDENCIES - OFFICIAL SOURCE ONLY
-----------------------------------------------
All Python packages this script depends on are installed via `pip install
<package>` with NO custom index/mirror configured. By default this pulls
exclusively from PyPI (pypi.org) - the Python Software Foundation's official
Python Package Index - and nowhere else. Do not run these installs with a
custom --index-url / -i pointing at a third-party or unofficial mirror.

    pip install pandas openpyxl          # required (CSV/.xlsx handling)
    pip install xlrd                     # optional, only if reading legacy .xls
    pip install pyxlsb                   # optional, only if reading .xlsb
    pip install anthropic                # optional, only if using --ai flag

You can verify any package's official PyPI page before installing by
visiting https://pypi.org/project/<package-name>/

REGULATORY SOURCES FOR THE DISCLOSURE LOGIC BUILT INTO THIS TOOL
  - Ind AS 107, para 35H (MCA notified text): requires, by class of financial
    instrument, a reconciliation from opening to closing loss allowance,
    showing changes separately for (a) 12-month ECL, (b)(i) lifetime ECL -
    not credit-impaired, (b)(ii) lifetime ECL - credit-impaired.
    Source: https://www.mca.gov.in/Ministry/pdf/IndAS107_2020_10112020.pdf
  - RBI (Non-Banking Financial Companies - Income Recognition, Asset
    Classification and Provisioning) Directions, 2025 (effective 28-Nov-2025):
    requires NBFCs on Ind AS to also compute provisioning under IRACP norms in
    parallel, disclose a comparison between the two, and appropriate any
    shortfall of Ind AS 109 below the IRACP floor to a separate 'Impairment
    Reserve' not reckoned for regulatory capital.
    Source: RBI IRACP Directions, 2025 (rbi.org.in)
  Always verify against the latest RBI/ICAI notifications before relying on
  this for an actual filing - this tool drafts the computation and layout,
  it does not substitute professional / regulatory judgement.

DPDPA / DATA PRIVACY NOTE
--------------------------
Everything in this script runs 100% locally - no data leaves your machine
unless you pass --ai, and even then ONLY aggregated stage totals (never
customer names or loan IDs) are sent to the AI API.
"""

import os
import re
import sys
import shutil
import hashlib
import argparse
from datetime import datetime
from itertools import product

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.worksheet.filters import AutoFilter

# Folder containing this script/executable. All default paths are anchored
# here so the tool behaves identically whether launched from IDLE, a
# double-click, a scheduled task, a terminal in some other directory, OR a
# PyInstaller-frozen .exe.
#
# Why the frozen check matters: PyInstaller's --onefile mode extracts the
# script (and its dependencies) into a TEMPORARY folder at runtime and sets
# __file__ to a path inside that temp folder, not to wherever the .exe
# actually sits. If SCRIPT_DIR were computed from __file__ unconditionally,
# a packaged .exe would create/read its data/, input_archive/ and output/
# folders inside a temp directory that gets wiped when the program exits -
# silently losing the user's data folder and every prior run's archive.
#
# sys.frozen is set to True by PyInstaller (and other freezers) on the
# running process; sys.executable then correctly points at the .exe itself,
# so anchoring to its folder keeps behaviour identical to the plain-script
# case: everything lives next to where the program is, not in a temp folder.
if getattr(sys, "frozen", False):
    SCRIPT_DIR = os.path.dirname(os.path.abspath(sys.executable))
else:
    SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))

# Windows consoles (and IDLE) often default to cp1252, which cannot encode
# characters such as the arrow used in period labels. Reconfigure stdout to
# UTF-8 where possible; ARROW falls back to a plain ASCII "->" if not.
try:
    sys.stdout.reconfigure(encoding="utf-8")
    sys.stderr.reconfigure(encoding="utf-8")
    ARROW = "\u2192"
except Exception:
    ARROW = "->"


def safe_print(*parts):
    """print() that degrades gracefully on a non-UTF-8 console."""
    text = " ".join(str(x) for x in parts)
    try:
        print(text)
    except UnicodeEncodeError:
        enc = sys.stdout.encoding or "ascii"
        print(text.encode(enc, errors="replace").decode(enc, errors="replace"))


def file_sha256(path: str) -> str:
    """SHA-256 of the input file, recorded in the run log for reproducibility."""
    h = hashlib.sha256()
    with open(path, "rb") as f:
        for chunk in iter(lambda: f.read(8192), b""):
            h.update(chunk)
    return h.hexdigest()


STAGES = ["Stage 1", "Stage 2", "Stage 3"]
STAGE_CODE = {"Stage 1": "S1", "Stage 2": "S2", "Stage 3": "S3"}

# --------------------------------------------------------------------------
# DPD (Days Past Due) BUCKETS
# --------------------------------------------------------------------------
# Standard ageing buckets used for NBFC delinquency reporting and SMA
# tracking. The 0 / 1-30 / 31-60 / 61-90 split aligns with RBI's SMA
# framework (SMA-0: 1-30 days, SMA-1: 31-60, SMA-2: 61-90), with the
# post-90 buckets reflecting NPA territory for an NBFC-Middle Layer
# (NPA at >90 days overdue, para 51 of the IRACP Directions, 2025).
DPD_BUCKETS = [
    ("0 (Current)", 0, 0),
    ("1-30", 1, 30),
    ("31-60", 31, 60),
    ("61-90", 61, 90),
    ("91-120", 91, 120),
    ("120+", 121, None),
]
DPD_BUCKET_NAMES = [b[0] for b in DPD_BUCKETS]


def assign_dpd_bucket(dpd) -> str:
    """Maps a DPD value to its ageing bucket label."""
    try:
        d = int(dpd)
    except (TypeError, ValueError):
        return "0 (Current)"
    if d < 0:
        d = 0
    for label, lo, hi in DPD_BUCKETS:
        if hi is None:
            if d >= lo:
                return label
        elif lo <= d <= hi:
            return label
    return DPD_BUCKET_NAMES[-1]

HEADER_FILL = PatternFill(start_color="FF1F4E78", end_color="FF1F4E78", fill_type="solid")
HEADER_FONT = Font(color="FFFFFFFF", bold=True)
DIAG_FILL = PatternFill(start_color="FFD9E1F2", end_color="FFD9E1F2", fill_type="solid")
WORSEN_FILL = PatternFill(start_color="FFFCE4E4", end_color="FFFCE4E4", fill_type="solid")
IMPROVE_FILL = PatternFill(start_color="FFE2EFDA", end_color="FFE2EFDA", fill_type="solid")
THIN_BORDER = Border(*(Side(style="thin", color="FFB7B7B7"),) * 4)
LINK_FONT = Font(color="FF0563C1", underline="single")
TITLE_FONT = Font(bold=True, size=13)
SUBTITLE_FONT = Font(italic=True, size=9, color="FF666666")


# --------------------------------------------------------------------------
# STEP 1: Load and validate loan-level input
# --------------------------------------------------------------------------
def load_data(filepath: str, require_period: bool = True, minimal: bool = False) -> pd.DataFrame:
    """
    Reads CSV or any common Excel format: .xlsx, .xls, .xlsm, .xlsb.

    Requires (only if that specific format is used):
        .xlsx / .xlsm  -> openpyxl        (pip install openpyxl)
        .xls  (legacy) -> xlrd            (pip install xlrd)
        .xlsb          -> pyxlsb          (pip install pyxlsb)
    """
    ext = os.path.splitext(filepath)[1].lower()

    if ext == ".csv":
        df = pd.read_csv(filepath)
    elif ext == ".xlsb":
        try:
            df = pd.read_excel(filepath, engine="pyxlsb")
        except ImportError:
            raise ImportError(
                "Reading .xlsb files requires the 'pyxlsb' package. "
                "Install it with:  pip install pyxlsb"
            )
    elif ext == ".xls":
        try:
            df = pd.read_excel(filepath, engine="xlrd")
        except ImportError:
            raise ImportError(
                "Reading legacy .xls files requires the 'xlrd' package. "
                "Install it with:  pip install xlrd"
            )
    elif ext in (".xlsx", ".xlsm"):
        df = pd.read_excel(filepath, engine="openpyxl")
    else:
        raise ValueError(
            f"Unsupported file type '{ext}'. Supported: .csv, .xlsx, .xlsm, .xls, .xlsb"
        )

    df.columns = [c.strip() for c in df.columns]
    if minimal:
        # write-off files need only a LoanID; everything else is optional
        return df
    required = {"Period", "LoanID", "CustomerName", "Stage", "Outstanding_Cr", "Provision_Cr"}
    if not require_period:
        # during a repository scan the period may come from the file name
        required = required - {"Period"}
    missing = required - set(df.columns)
    if missing:
        raise ValueError(f"missing required columns: {sorted(missing)}")

    df["Stage"] = df["Stage"].str.strip()
    df["Period"] = df["Period"].astype(str).str.strip()
    return df


HAS_IRACP_COL = "IRACP_Provision_Cr"

# --------------------------------------------------------------------------
# IRACP BUCKET-WISE AUTO-CLASSIFICATION & PROVISIONING  --  NBFC-MIDDLE LAYER
# --------------------------------------------------------------------------
# SCOPE: This tool implements the IRACP norms for an NBFC classified in the
# MIDDLE LAYER (NBFC-ML) ONLY. Base Layer and Upper Layer rates are NOT
# implemented, by design. If your NBFC is re-classified into another layer,
# the constants below must be revisited before the output is relied upon.
#
# Source: Reserve Bank of India (Non-Banking Financial Companies - Income
# Recognition, Asset Classification and Provisioning) Directions, 2025
# RBI/DOR/2025-26/356 dated November 28, 2025 (effective immediately).
#
#   NPA threshold - para 51:
#       An asset is an NPA where interest / instalment is overdue for more
#       than 90 days.
#
#   Sub-standard / Doubtful boundary - para 52-53:
#       Sub-standard : an asset which has been an NPA for a period not
#                      exceeding 12 months.
#       Doubtful     : an asset which remains sub-standard for a period
#                      exceeding 12 months.
#
#   Provisioning - para 32 (all NBFCs) read with para 55 (NBFC-ML):
#       Standard     : 0.40 per cent of outstanding                [para 55]
#       Sub-standard : 10 per cent of total outstanding         [para 32(1)]
#       Doubtful     : 100 per cent of the portion NOT covered by the
#                      realisable value of security, PLUS, on the secured
#                      portion, by period as doubtful:          [para 32(2)]
#                          up to 1 year        -> 20 per cent
#                          1 to 3 years        -> 30 per cent
#                          more than 3 years   -> 50 per cent
#       Loss         : 100 per cent of outstanding              [para 32(3)]
#                      A loss asset is one identified as such by the NBFC,
#                      its internal / external auditor, or the Reserve Bank
#                      on inspection [para 11(1)]. It CANNOT be derived from
#                      DPD alone - it must be flagged via the Loss_Flag
#                      column. Absent that flag, no asset is treated as loss.
#
# NOT IMPLEMENTED (state these in the write-up; see the auto-generated
# "Limitations" sheet): project finance provisioning [para 30], securitisation
# liquidity facilities [para 31], hire-purchase / lease-specific norms
# [para 33], and restructured / DCCO accounts governed by the RBI (NBFC -
# Resolution of Stressed Assets) Directions, 2025 [para 9, 22, 26].

NPA_THRESHOLD_DAYS = 90          # para 51
SUBSTANDARD_PERIOD_DAYS = 365    # para 52-53 : 12 months as NPA
STANDARD_PROVISION_RATE = 0.0040 # para 55    : NBFC-ML, 0.40%
SUBSTANDARD_PROVISION_RATE = 0.10  # para 32(1)
DOUBTFUL_SECURED_RATES = [       # para 32(2)
    (365, 0.20),                 # doubtful up to 1 year
    (1095, 0.30),                # doubtful 1 to 3 years
    (None, 0.50),                # doubtful more than 3 years
]


def classify_iracp_bucket(dpd: int):
    """
    Classifies one loan per RBI IRACP Directions, 2025 (NBFC-ML).
    Returns (asset_category, days_as_doubtful_or_None).
    """
    if dpd <= NPA_THRESHOLD_DAYS:               # para 51
        return "Standard", None
    days_as_npa = dpd - NPA_THRESHOLD_DAYS
    if days_as_npa <= SUBSTANDARD_PERIOD_DAYS:  # para 52
        return "Sub-standard", None
    return "Doubtful", days_as_npa - SUBSTANDARD_PERIOD_DAYS   # para 53


def compute_iracp_provision_for_loan(outstanding_cr: float, dpd: int, loss_flag: bool,
                                      secured_value_cr=None):
    """
    Returns a dict with the FULL working for one loan, not just the final
    provision number - so the loan-level sheet can show the secured/
    unsecured split and the rate(s) applied, and write the provision itself
    as a genuine Excel formula rather than a value baked in by Python.

    The four figures below (secured_cr, unsecured_cr, rate_secured_pct,
    rate_unsecured_pct) are defined so that, for EVERY category, the RBI
    provision is always:

        provision = secured_cr * rate_secured_pct + unsecured_cr * rate_unsecured_pct

    For Standard / Sub-standard / Loss - where para 32/55 apply a single
    flat rate to the WHOLE outstanding regardless of security - this is
    modelled by putting the entire outstanding in "unsecured_cr" at that
    flat rate, with "secured_cr" at zero. This is a presentational
    convenience so one formula works everywhere; it does not change the
    computed amount, and the sheet's "Basis" column states in words which
    rule was actually applied for that row.

    SECURITY DATA IS OPTIONAL for Doubtful assets. Where Secured_Value_Cr is
    absent, blank, or non-numeric, the loan is treated as FULLY UNSECURED -
    prudent, since para 32(2)(i) requires 100% on the uncovered portion, so
    this can only OVERSTATE the floor, never understate it. Flagged on the
    Assumptions sheet.

    Keys returned:
        category            : "Standard" / "Sub-standard" / "Doubtful" / "Loss"
        secured_cr          : amount attracting rate_secured_pct
        unsecured_cr        : amount attracting rate_unsecured_pct
        rate_secured_pct    : rate applied to secured_cr (0-1, e.g. 0.30)
        rate_unsecured_pct  : rate applied to unsecured_cr (0-1)
        provision_cr        : secured_cr*rate_secured_pct + unsecured_cr*rate_unsecured_pct
        basis               : human-readable description of the rule applied
        security_basis      : "actual" / "assumed-nil" / "n/a"
    """
    if loss_flag:
        prov = outstanding_cr * 1.00
        return {"category": "Loss", "secured_cr": 0.0, "unsecured_cr": outstanding_cr,
                "rate_secured_pct": 0.0, "rate_unsecured_pct": 1.00, "provision_cr": prov,
                "basis": "Loss asset - 100% of outstanding [para 32(3)]", "security_basis": "n/a"}

    category, doubtful_days = classify_iracp_bucket(dpd)

    if category == "Standard":
        prov = outstanding_cr * STANDARD_PROVISION_RATE
        return {"category": "Standard", "secured_cr": 0.0, "unsecured_cr": outstanding_cr,
                "rate_secured_pct": 0.0, "rate_unsecured_pct": STANDARD_PROVISION_RATE,
                "provision_cr": prov,
                "basis": f"Standard asset, NBFC-ML - {STANDARD_PROVISION_RATE:.2%} flat [para 55]",
                "security_basis": "n/a"}

    if category == "Sub-standard":
        prov = outstanding_cr * SUBSTANDARD_PROVISION_RATE
        return {"category": "Sub-standard", "secured_cr": 0.0, "unsecured_cr": outstanding_cr,
                "rate_secured_pct": 0.0, "rate_unsecured_pct": SUBSTANDARD_PROVISION_RATE,
                "provision_cr": prov,
                "basis": f"Sub-standard - {SUBSTANDARD_PROVISION_RATE:.0%} of total outstanding [para 32(1)]",
                "security_basis": "n/a"}

    # --- Doubtful: para 32(2) ---
    if secured_value_cr is None or (isinstance(secured_value_cr, float) and pd.isna(secured_value_cr)):
        secured, basis_tag = 0.0, "assumed-nil"
    else:
        try:
            secured = min(max(float(secured_value_cr), 0.0), outstanding_cr)
            basis_tag = "actual"
        except (TypeError, ValueError):
            secured, basis_tag = 0.0, "assumed-nil"

    unsecured = outstanding_cr - secured
    for threshold, rate in DOUBTFUL_SECURED_RATES:
        if threshold is None or doubtful_days <= threshold:
            secured_rate = rate
            age_label = ("up to 1 year" if threshold == 365 else
                         "1 to 3 years" if threshold == 1095 else "more than 3 years")
            break
    prov = unsecured * 1.00 + secured * secured_rate
    return {"category": "Doubtful", "secured_cr": secured, "unsecured_cr": unsecured,
            "rate_secured_pct": secured_rate, "rate_unsecured_pct": 1.00, "provision_cr": prov,
            "basis": (f"Doubtful, {age_label} - 100% of unsecured + {secured_rate:.0%} "
                      f"of secured [para 32(2)]"),
            "security_basis": basis_tag}


# --------------------------------------------------------------------------
# STATUTORY DISCLOSURE 1: Ind AS 107 para 35H - loss allowance reconciliation
# --------------------------------------------------------------------------
# Ind AS 107 para 35H categories map to our stages as follows for a simple
# NBFC retail/SME book (state this mapping assumption explicitly in your
# capstone write-up, since the actual Ind AS category depends on your
# significant-increase-in-credit-risk (SICR) policy, not just DPD buckets):
#   12-month ECL                          <- Stage 1
#   Lifetime ECL - not credit-impaired    <- Stage 2
#   Lifetime ECL - credit-impaired        <- Stage 3
MOVEMENT_TYPES = [
    "Opening balance",
    "New assets originated",
    "Amounts written off",
    "Assets derecognised / repaid",
    "Transferred out to other stage",
    "Transferred in from other stage",
    "Net remeasurement (no stage change)",
    "Closing balance",
]

# --------------------------------------------------------------------------
# WRITE-OFFS
# --------------------------------------------------------------------------
# Ind AS 107 para 35H requires the loss-allowance reconciliation to show
# movements separately, and para 35I(c)/35L call for write-offs to be
# disclosed distinctly - a write-off reduces the allowance without any cash
# recovery, so it must NOT be lumped in with ordinary repayment.
#
# Two ways to tell the tool a loan was written off (either or both):
#   1. TAG IN THE PERIOD FILE  - a WriteOff_Flag column ("Y"/"N"), and
#      optionally WriteOff_Amount_Cr. A tagged loan usually then disappears
#      from the following period file.
#   2. SEPARATE WRITE-OFF FILE - a file in the same data folder whose name
#      contains "writeoff" / "write_off" / "write-off" / "wo_", e.g.
#          WriteOff_Q3FY26.csv     write_off_Q3FY26.xlsx
#      containing at least LoanID, and optionally Period, CustomerName,
#      WriteOff_Amount_Cr. If Period is absent it is taken from the file name.
#
# Where an amount is not supplied, the loan's provision held in the period it
# was last seen is used as the amount written off (the prudent reading: the
# allowance built against that asset is what is released on write-off).
WRITEOFF_FILE_MARKERS = ("writeoff", "write_off", "write-off", "write off", "wo_", "_wo", "w_off", "w-off")


def is_writeoff_file(filename: str) -> bool:
    stem = os.path.splitext(os.path.basename(filename))[0].lower()
    return any(m in stem for m in WRITEOFF_FILE_MARKERS)


def extract_period_from_name(filename: str, known_periods) -> str:
    """Finds which known period a write-off file name refers to."""
    stem = os.path.splitext(os.path.basename(filename))[0].upper()
    for per in known_periods:
        if per.upper() in stem:
            return per
    return None


def collect_writeoffs(df: pd.DataFrame, writeoff_frames: dict, period: str) -> pd.DataFrame:
    """
    Returns the write-offs applicable to `period`, from the in-file tag and/or
    a separate write-off file, de-duplicated by LoanID.
    Columns: LoanID, CustomerName, Stage, WriteOff_Amount_Cr, Source
    """
    rows = []

    pdf = df[df["Period"] == period]
    if "WriteOff_Flag" in pdf.columns:
        tagged = pdf[pdf["WriteOff_Flag"].astype(str).str.strip().str.upper() == "Y"]
        for _, r in tagged.iterrows():
            amt = r.get("WriteOff_Amount_Cr", None)
            try:
                amt = float(amt)
                if pd.isna(amt):
                    raise ValueError
            except (TypeError, ValueError):
                amt = float(r.get("Provision_Cr", 0) or 0)
            rows.append({"LoanID": r["LoanID"], "CustomerName": r.get("CustomerName", ""),
                          "Stage": r.get("Stage", ""), "WriteOff_Amount_Cr": round(amt, 3),
                          "Source": "Tagged in period file"})

    wo = writeoff_frames.get(period)
    if wo is not None:
        for _, r in wo.iterrows():
            lid = r["LoanID"]
            amt = r.get("WriteOff_Amount_Cr", None)
            try:
                amt = float(amt)
                if pd.isna(amt):
                    raise ValueError
            except (TypeError, ValueError):
                match = pdf[pdf["LoanID"] == lid]
                amt = float(match["Provision_Cr"].iloc[0]) if len(match) else 0.0
            match = pdf[pdf["LoanID"] == lid]
            rows.append({
                "LoanID": lid,
                "CustomerName": r.get("CustomerName", match["CustomerName"].iloc[0] if len(match) else ""),
                "Stage": match["Stage"].iloc[0] if len(match) else r.get("Stage", "Stage 3"),
                "WriteOff_Amount_Cr": round(amt, 3),
                "Source": "Separate write-off file",
            })

    if not rows:
        return pd.DataFrame(columns=["LoanID", "CustomerName", "Stage",
                                      "WriteOff_Amount_Cr", "Source"])
    out = pd.DataFrame(rows).drop_duplicates(subset="LoanID", keep="first")
    return out.reset_index(drop=True)


def compute_indas107_reconciliation(df: pd.DataFrame, period_from: str, period_to: str,
                                     writeoffs: pd.DataFrame = None) -> pd.DataFrame:
    df_from = df[df["Period"] == period_from][["LoanID", "CustomerName", "Stage", "Provision_Cr"]]
    df_to = df[df["Period"] == period_to][["LoanID", "CustomerName", "Stage", "Provision_Cr"]]

    from_ids = set(df_from["LoanID"])
    to_ids = set(df_to["LoanID"])

    closed = df_from[~df_from["LoanID"].isin(to_ids)]
    new = df_to[~df_to["LoanID"].isin(from_ids)]
    merged = df_from.merge(df_to, on=["LoanID", "CustomerName"], suffixes=("_from", "_to"))

    data = {}
    for stage in STAGES:
        opening = df_from.loc[df_from["Stage"] == stage, "Provision_Cr"].sum()
        closing = df_to.loc[df_to["Stage"] == stage, "Provision_Cr"].sum()

        transferred_out = merged.loc[
            (merged["Stage_from"] == stage) & (merged["Stage_to"] != stage), "Provision_Cr_from"
        ].sum()
        transferred_in = merged.loc[
            (merged["Stage_to"] == stage) & (merged["Stage_from"] != stage), "Provision_Cr_to"
        ].sum()

        # Split exits into write-offs (no cash recovery - disclosed
        # separately per Ind AS 107) and ordinary repayment/derecognition.
        #
        # IMPORTANT: the write-off amount used HERE must come from the same
        # basis as every other line in this reconciliation - the OPENING
        # period's (period_from) stage and provision, taken from `closed`
        # (which is itself built from df_from). Using the write-off
        # registry's own amount (which may be sourced from the write-off
        # period's raw, pre-exclusion data, or an explicit override amount)
        # would mix two different valuation bases and break the identity
        # opening + movements = closing. The registry's own amount is still
        # the right figure for the commentary and the migration sheet's
        # entries/exits line, since that reflects the actual write-off
        # transaction - it is only THIS internal reconciliation that needs
        # the opening-period figure to tie out arithmetically.
        wo_ids = set(writeoffs["LoanID"]) if writeoffs is not None and len(writeoffs) else set()

        exited = closed[closed["Stage"] == stage]
        wo_exited = exited[exited["LoanID"].isin(wo_ids)]
        non_wo_exited = exited[~exited["LoanID"].isin(wo_ids)]
        wo_amt = wo_exited["Provision_Cr"].sum()
        derecognised = non_wo_exited["Provision_Cr"].sum()
        originated = new.loc[new["Stage"] == stage, "Provision_Cr"].sum()

        stayed = merged[(merged["Stage_from"] == stage) & (merged["Stage_to"] == stage)]
        remeasurement = (stayed["Provision_Cr_to"] - stayed["Provision_Cr_from"]).sum()

        data[stage] = [
            opening, originated, -wo_amt, -derecognised, -transferred_out,
            transferred_in, remeasurement, closing,
        ]

    result = pd.DataFrame(data, index=MOVEMENT_TYPES)
    result["Total"] = result.sum(axis=1)
    return result.round(3)


# --------------------------------------------------------------------------
# STATUTORY DISCLOSURE 2: RBI IRACP vs Ind AS 109 comparison (NBFC-specific)
# --------------------------------------------------------------------------
IRACP_CATEGORIES = ["Standard", "Sub-standard", "Doubtful", "Loss"]


def compute_iracp_comparison(df: pd.DataFrame, period: str):
    """
    Auto-computes IRACP asset category and provisioning PER LOAN for an
    NBFC-MIDDLE LAYER, from bucket-wise data. Returns
    (comparison_table, method_note, assumptions_table, loan_level_detail).

    MINIMUM REQUIREMENT: a DPD column. That alone is enough to run.
      - Secured_Value_Cr : OPTIONAL. Missing/blank values are treated as nil
                           security (prudent - see compute_iracp_provision_for_loan).
      - Loss_Flag        : OPTIONAL. Absent means no asset is classified as a
                           loss asset, which is correct: para 11(1) requires
                           positive identification by the NBFC, its auditor or
                           the Reserve Bank; it is not derivable from DPD.
    """
    pdf = df[df["Period"] == period].copy()

    if "DPD" not in pdf.columns:
        return None, None, None, None

    has_security = "Secured_Value_Cr" in pdf.columns
    has_loss_flag = "Loss_Flag" in pdf.columns

    working_rows = []
    for _, row in pdf.iterrows():
        loss_flag = (str(row.get("Loss_Flag", "N")).strip().upper() == "Y") if has_loss_flag else False
        sec_val = row.get("Secured_Value_Cr", None) if has_security else None
        w = compute_iracp_provision_for_loan(
            outstanding_cr=float(row["Outstanding_Cr"]),
            dpd=int(row["DPD"]),
            loss_flag=loss_flag,
            secured_value_cr=sec_val,
        )
        w["LoanID"] = row["LoanID"]
        working_rows.append(w)

    work_df = pd.DataFrame(working_rows)
    pdf = pdf.reset_index(drop=True)
    pdf["IRACP_Category"] = work_df["category"]
    pdf["Secured_Cr"] = work_df["secured_cr"].round(4)
    pdf["Unsecured_Cr"] = work_df["unsecured_cr"].round(4)
    pdf["Rate_Secured_pct"] = work_df["rate_secured_pct"]
    pdf["Rate_Unsecured_pct"] = work_df["rate_unsecured_pct"]
    pdf["IRACP_Provision_Cr"] = work_df["provision_cr"].round(4)
    pdf["Basis"] = work_df["basis"]
    pdf["Security_Basis"] = work_df["security_basis"]

    notes = ["Auto-computed per loan from DPD. NBFC-Middle Layer rates "
             "(RBI IRACP Directions, 2025 - para 51/52/53/32/55)."]
    if not has_security:
        notes.append("No Secured_Value_Cr column supplied: all doubtful assets treated as "
                     "FULLY UNSECURED (prudent - overstates the floor, never understates it).")
    if not has_loss_flag:
        notes.append("No Loss_Flag column supplied: no asset classified as a loss asset "
                     "(para 11(1) requires positive identification, not derivable from DPD).")
    method_note = " ".join(notes)

    # ---- Assumptions register: every loan where security was assumed nil ----
    assumed = pdf[pdf["Security_Basis"] == "assumed-nil"]
    if len(assumed):
        assumptions_table = pd.DataFrame([{
            "Assumption": "Doubtful assets treated as fully unsecured (no realisable security value supplied)",
            "Regulatory basis": "Para 32(2)(i) - 100% provision on portion not covered by realisable security",
            "Effect on IRACP floor": "Overstates (prudent). Never understates.",
            "Loans affected": assumed["LoanID"].nunique(),
            "Outstanding affected (Rs. Cr)": round(assumed["Outstanding_Cr"].sum(), 2),
            "IRACP provision on these loans (Rs. Cr)": round(assumed["IRACP_Provision_Cr"].sum(), 3),
        }])
    else:
        assumptions_table = pd.DataFrame([{
            "Assumption": "None - realisable security value was supplied for every doubtful asset",
            "Regulatory basis": "Para 32(2)",
            "Effect on IRACP floor": "Nil",
            "Loans affected": 0,
            "Outstanding affected (Rs. Cr)": 0.0,
            "IRACP provision on these loans (Rs. Cr)": 0.0,
        }])

    # loan_detail carries everything needed to write LIVE FORMULAS on the
    # sheet: Secured_Cr, Unsecured_Cr, and both rates are the actual inputs;
    # the provision itself is written as an Excel formula referencing them
    # (see write_iracp_loan_working_sheet), not as a value computed only in
    # Python - so anyone opening the file can click the cell and see exactly
    # how the number was derived.
    loan_detail = pdf[["LoanID", "CustomerName", "Stage", "DPD", "IRACP_Category",
                        "Outstanding_Cr", "Secured_Cr", "Unsecured_Cr",
                        "Rate_Secured_pct", "Rate_Unsecured_pct",
                        "IRACP_Provision_Cr", "Provision_Cr", "Security_Basis", "Basis"]].copy()
    loan_detail = loan_detail.rename(columns={
        "Stage": "Ind AS 109 Stage",
        "Provision_Cr": "Ind AS 109 ECL (Rs. Cr)",
        "IRACP_Provision_Cr": "IRACP Provision (Rs. Cr)",
    })

    categories_order = IRACP_CATEGORIES
    rate_summary = {
        "Standard": f"{STANDARD_PROVISION_RATE:.2%} of outstanding [para 55]",
        "Sub-standard": f"{SUBSTANDARD_PROVISION_RATE:.0%} of outstanding [para 32(1)]",
        "Doubtful": "100% of unsecured + 20%/30%/50% of secured by ageing [para 32(2)]",
        "Loss": "100% of outstanding [para 32(3)]",
    }

    rows = []
    total_indas, total_iracp = 0.0, 0.0
    for cat in categories_order:
        cdf = pdf[pdf["IRACP_Category"] == cat]
        if cdf.empty:
            continue
        indas_prov = cdf["Provision_Cr"].sum()
        iracp_prov = cdf["IRACP_Provision_Cr"].sum()
        rows.append({
            "IRACP Asset Category": cat,
            "Rate(s) Applied": rate_summary[cat],
            "Loan_Count": cdf["LoanID"].nunique(),
            "Outstanding_Cr": round(cdf["Outstanding_Cr"].sum(), 2),
            "Secured_Cr": round(cdf["Secured_Cr"].sum(), 2),
            "Unsecured_Cr": round(cdf["Unsecured_Cr"].sum(), 2),
            "Ind_AS_109_ECL_Provision_Cr": round(indas_prov, 3),
            "IRACP_Provision_Cr": round(iracp_prov, 3),
            "Excess/(Shortfall) of Ind AS 109 over IRACP_Cr": round(indas_prov - iracp_prov, 3),
        })
        total_indas += indas_prov
        total_iracp += iracp_prov

    rows.append({
        "IRACP Asset Category": "TOTAL",
        "Rate(s) Applied": "",
        "Loan_Count": pdf["LoanID"].nunique(),
        "Outstanding_Cr": round(pdf["Outstanding_Cr"].sum(), 2),
        "Secured_Cr": round(pdf["Secured_Cr"].sum(), 2),
        "Unsecured_Cr": round(pdf["Unsecured_Cr"].sum(), 2),
        "Ind_AS_109_ECL_Provision_Cr": round(total_indas, 3),
        "IRACP_Provision_Cr": round(total_iracp, 3),
        "Excess/(Shortfall) of Ind AS 109 over IRACP_Cr": round(total_indas - total_iracp, 3),
    })
    return pd.DataFrame(rows), method_note, assumptions_table, loan_detail


# --------------------------------------------------------------------------
# STEP 2: Period-wise aggregate summary (rolled up from loan-level data)
# --------------------------------------------------------------------------
def compute_period_summary(df: pd.DataFrame) -> pd.DataFrame:
    periods = list(dict.fromkeys(df["Period"]))
    rows = []
    for period in periods:
        pdf = df[df["Period"] == period]
        total_out = pdf["Outstanding_Cr"].sum()
        total_prov = pdf["Provision_Cr"].sum()
        row = {
            "Period": period,
            "Loan_Count": pdf["LoanID"].nunique(),
            "Total_Outstanding_Cr": round(total_out, 2),
            "Total_Provision_Cr": round(total_prov, 3),
            "Overall_PCR_pct": round(100 * total_prov / total_out, 2) if total_out else 0,
        }
        for stage in STAGES:
            sdf = pdf[pdf["Stage"] == stage]
            out = sdf["Outstanding_Cr"].sum()
            prov = sdf["Provision_Cr"].sum()
            row[f"{stage}_Count"] = sdf["LoanID"].nunique()
            row[f"{stage}_Outstanding_Cr"] = round(out, 2)
            row[f"{stage}_Provision_Cr"] = round(prov, 3)
            row[f"{stage}_Share_pct"] = round(100 * out / total_out, 2) if total_out else 0
        rows.append(row)
    return pd.DataFrame(rows)


# --------------------------------------------------------------------------
# STEP 3: Build From-Stage x To-Stage migration matrix for a period pair
# --------------------------------------------------------------------------
def build_migration_matrix(df: pd.DataFrame, period_from: str, period_to: str):
    """
    Returns:
        matrix_count : DataFrame, index=From Stage, columns=To Stage -> loan count
        matrix_amount: DataFrame, same shape -> outstanding Rs. Cr (as at period_to)
        detail       : dict{(from_stage, to_stage): DataFrame of loan-level rows}
    Only loans present in BOTH periods are considered "migrating" (or "static").
    Loans only in period_from = closed/paid off. Loans only in period_to = new.

    If the input carries Product and/or Branch columns, they are carried
    through into `detail` (and therefore into the consolidated drill-down)
    so the loan-level view can be filtered by either without touching the
    stage/DPD matrices themselves.
    """
    extra_cols = [c for c in ("Product", "Branch") if c in df.columns]
    base_cols = ["LoanID", "CustomerName", "Stage", "Outstanding_Cr", "Provision_Cr"] + extra_cols

    df_from = df[df["Period"] == period_from][base_cols]
    df_to = df[df["Period"] == period_to][base_cols]

    merge_keys = ["LoanID", "CustomerName"] + extra_cols
    merged = df_from.merge(
        df_to, on=merge_keys, suffixes=("_from", "_to"), how="inner"
    )

    matrix_count = pd.DataFrame(0, index=STAGES, columns=STAGES)
    matrix_amount = pd.DataFrame(0.0, index=STAGES, columns=STAGES)
    matrix_prov = pd.DataFrame(0.0, index=STAGES, columns=STAGES)
    detail = {}

    detail_cols = ["LoanID", "CustomerName"] + extra_cols + [
        "Outstanding_Cr_from", "Outstanding_Cr_to", "Provision_Cr_from", "Provision_Cr_to"
    ]
    rename_map = {
        "Outstanding_Cr_from": f"Outstanding_Cr ({period_from})",
        "Outstanding_Cr_to": f"Outstanding_Cr ({period_to})",
        "Provision_Cr_from": f"Provision_Cr ({period_from})",
        "Provision_Cr_to": f"Provision_Cr ({period_to})",
    }

    for from_stage, to_stage in product(STAGES, STAGES):
        cell = merged[(merged["Stage_from"] == from_stage) & (merged["Stage_to"] == to_stage)]
        matrix_count.loc[from_stage, to_stage] = len(cell)
        matrix_amount.loc[from_stage, to_stage] = round(cell["Outstanding_Cr_to"].sum(), 2)
        matrix_prov.loc[from_stage, to_stage] = round(cell["Provision_Cr_to"].sum(), 3)
        detail[(from_stage, to_stage)] = cell[detail_cols].rename(columns=rename_map)

    closed = df_from[~df_from["LoanID"].isin(df_to["LoanID"])]
    new = df_to[~df_to["LoanID"].isin(df_from["LoanID"])]

    return matrix_count, matrix_amount, matrix_prov, detail, closed, new


def build_dpd_migration(df: pd.DataFrame, period_from: str, period_to: str):
    """
    Builds a DPD-bucket migration matrix: rows = ageing bucket in the earlier
    period, columns = ageing bucket in the later period. Buckets are
    0 (Current) / 1-30 / 31-60 / 61-90 / 91-120 / 120+.

    Returns (matrix_count, matrix_amount, flat_detail_df) or None if the
    input has no DPD column.
    """
    if "DPD" not in df.columns:
        return None

    cols = ["LoanID", "CustomerName", "DPD", "Outstanding_Cr", "Provision_Cr", "Stage"]
    df_from = df[df["Period"] == period_from][cols].copy()
    df_to = df[df["Period"] == period_to][cols].copy()
    df_from["Bucket"] = df_from["DPD"].apply(assign_dpd_bucket)
    df_to["Bucket"] = df_to["DPD"].apply(assign_dpd_bucket)

    merged = df_from.merge(df_to, on=["LoanID", "CustomerName"],
                            suffixes=("_from", "_to"), how="inner")

    matrix_count = pd.DataFrame(0, index=DPD_BUCKET_NAMES, columns=DPD_BUCKET_NAMES)
    matrix_amount = pd.DataFrame(0.0, index=DPD_BUCKET_NAMES, columns=DPD_BUCKET_NAMES)
    matrix_prov = pd.DataFrame(0.0, index=DPD_BUCKET_NAMES, columns=DPD_BUCKET_NAMES)

    for b_from, b_to in product(DPD_BUCKET_NAMES, DPD_BUCKET_NAMES):
        cell = merged[(merged["Bucket_from"] == b_from) & (merged["Bucket_to"] == b_to)]
        matrix_count.loc[b_from, b_to] = len(cell)
        matrix_amount.loc[b_from, b_to] = round(cell["Outstanding_Cr_to"].sum(), 2)
        matrix_prov.loc[b_from, b_to] = round(cell["Provision_Cr_to"].sum(), 3)

    flat = merged[["LoanID", "CustomerName", "Bucket_from", "Bucket_to",
                    "DPD_from", "DPD_to", "Stage_from", "Stage_to",
                    "Outstanding_Cr_from", "Outstanding_Cr_to"]].copy()
    flat["Movement"] = [
        "Deteriorated" if DPD_BUCKET_NAMES.index(r["Bucket_to"]) > DPD_BUCKET_NAMES.index(r["Bucket_from"])
        else "Improved" if DPD_BUCKET_NAMES.index(r["Bucket_to"]) < DPD_BUCKET_NAMES.index(r["Bucket_from"])
        else "No change"
        for _, r in flat.iterrows()
    ]
    flat = flat.rename(columns={
        "Bucket_from": f"DPD Bucket ({period_from})",
        "Bucket_to": f"DPD Bucket ({period_to})",
        "DPD_from": f"DPD ({period_from})",
        "DPD_to": f"DPD ({period_to})",
        "Stage_from": f"Stage ({period_from})",
        "Stage_to": f"Stage ({period_to})",
        "Outstanding_Cr_from": f"Outstanding Cr ({period_from})",
        "Outstanding_Cr_to": f"Outstanding Cr ({period_to})",
    })
    flat = flat.sort_values("Movement")
    return matrix_count, matrix_amount, matrix_prov, flat


def build_group_summary(df: pd.DataFrame, period_from: str, period_to: str, group_col: str) -> pd.DataFrame:
    """
    Aggregates outstanding, provision, loan count and Stage 3 (credit-impaired)
    share by an arbitrary grouping column (Product or Branch), for the two
    selected periods, with period-on-period deltas. Used to build the
    Product Summary sheet - and reused for the filterable Branch sheet's
    "static" reference columns before the live SUBTOTAL rows are added.
    """
    out_rows = []
    for grp in sorted(df[group_col].dropna().unique()):
        row = {group_col: grp}
        for label, per in (("from", period_from), ("to", period_to)):
            gdf = df[(df["Period"] == per) & (df[group_col] == grp)]
            total_out = gdf["Outstanding_Cr"].sum()
            total_prov = gdf["Provision_Cr"].sum()
            s3_out = gdf.loc[gdf["Stage"] == "Stage 3", "Outstanding_Cr"].sum()
            row[f"Loans_{label}"] = gdf["LoanID"].nunique()
            row[f"Outstanding_Cr_{label}"] = round(total_out, 2)
            row[f"Provision_Cr_{label}"] = round(total_prov, 3)
            row[f"Stage3_Share_pct_{label}"] = round(100 * s3_out / total_out, 2) if total_out else 0.0
        row["Δ Outstanding_Cr"] = round(row["Outstanding_Cr_to"] - row["Outstanding_Cr_from"], 2)
        row["Δ Provision_Cr"] = round(row["Provision_Cr_to"] - row["Provision_Cr_from"], 3)
        row["Δ Stage3 Share (pp)"] = round(row["Stage3_Share_pct_to"] - row["Stage3_Share_pct_from"], 2)
        out_rows.append(row)
    result = pd.DataFrame(out_rows)
    if not result.empty:
        result = result.sort_values(f"Outstanding_Cr_to", ascending=False).reset_index(drop=True)
    return result



def build_consolidated_drilldown(detail: dict, period_from: str, period_to: str) -> pd.DataFrame:
    """
    (instead of 9 separate sheets). Adds From/To/Movement columns so the
    user can filter in Excel rather than hopping between sheets.
    """
    frames = []
    for (from_stage, to_stage), ddf in detail.items():
        if ddf.empty:
            continue
        d = ddf.copy()
        d.insert(0, f"Stage ({period_from})", from_stage)
        d.insert(1, f"Stage ({period_to})", to_stage)
        if STAGES.index(to_stage) > STAGES.index(from_stage):
            mv = "Deteriorated"
        elif STAGES.index(to_stage) < STAGES.index(from_stage):
            mv = "Improved"
        else:
            mv = "No change"
        d.insert(2, "Movement", mv)
        frames.append(d)
    if not frames:
        return pd.DataFrame()
    out = pd.concat(frames, ignore_index=True)
    order = {"Deteriorated": 0, "Improved": 1, "No change": 2}
    out["_o"] = out["Movement"].map(order)
    return out.sort_values(["_o", "LoanID"]).drop(columns="_o").reset_index(drop=True)


# --------------------------------------------------------------------------
# STEP 4: Rule-based commentary (deterministic, fully offline)
# --------------------------------------------------------------------------
def generate_rule_based_commentary(summary: pd.DataFrame, migrations: list,
                                    pairs: list = None, writeoffs: pd.DataFrame = None,
                                    product_summary: pd.DataFrame = None) -> str:
    """
    Drafts commentary for the LAST comparison pair actually requested, not
    merely the last two periods in the file. So --compare Q1FY26:Q3FY26
    produces commentary comparing Q3 against Q1, with movement described
    over that span rather than quarter-on-quarter.
    """
    if pairs:
        period_from, period_to = pairs[-1]
        span_label = f"{period_from} {ARROW} {period_to}"
    else:
        period_from = summary.iloc[-2]["Period"] if len(summary) > 1 else None
        period_to = summary.iloc[-1]["Period"]
        span_label = f"{period_from} {ARROW} {period_to}" if period_from else period_to

    idx = {p: i for i, p in enumerate(summary["Period"])}
    latest = summary.iloc[idx[period_to]]
    prior = summary.iloc[idx[period_from]] if period_from in idx else None
    period = period_to

    lines = [f"ECL PROVISIONING COMMENTARY — {period}",
             f"Comparison basis: {span_label}\n"]

    # ---- 1. Headline position, in plain language ----
    lines.append("PORTFOLIO POSITION")
    lines.append(
        f"The book stood at Rs. {latest['Total_Outstanding_Cr']:,.2f} Cr across "
        f"{int(latest['Loan_Count'])} loans as at {period}. ECL provision held was "
        f"Rs. {latest['Total_Provision_Cr']:,.3f} Cr, giving a coverage of "
        f"{latest['Overall_PCR_pct']:.2f}% of the book."
    )
    if prior is not None:
        book_chg = latest["Total_Outstanding_Cr"] - prior["Total_Outstanding_Cr"]
        prov_chg = latest["Total_Provision_Cr"] - prior["Total_Provision_Cr"]
        pcr_chg = latest["Overall_PCR_pct"] - prior["Overall_PCR_pct"]
        lines.append(
            f"Against {period_from}, the book has "
            f"{'grown' if book_chg > 0 else 'reduced'} by Rs. {abs(book_chg):,.2f} Cr and "
            f"provision has {'risen' if prov_chg > 0 else 'fallen'} by Rs. {abs(prov_chg):,.3f} Cr. "
            f"Coverage {'improved' if pcr_chg > 0 else 'eased'} by {abs(pcr_chg):.2f} percentage points."
        )

    # ---- 2. Where the book sits, stage by stage ----
    lines.append("\nASSET QUALITY BY STAGE")
    for stage in STAGES:
        share = latest[f"{stage}_Share_pct"]
        cnt = int(latest[f"{stage}_Count"])
        amt = latest[f"{stage}_Outstanding_Cr"]
        line = f"{stage}: {cnt} loans, Rs. {amt:,.2f} Cr ({share:.2f}% of the book)"
        if prior is not None:
            delta = share - prior[f"{stage}_Share_pct"]
            if abs(delta) < 0.05:
                line += " — broadly unchanged."
            else:
                line += f" — {'up' if delta > 0 else 'down'} {abs(delta):.2f} pp."
        lines.append(line)

    # ---- 3. What actually moved, and what it means ----
    if migrations:
        latest_mig = migrations[-1]
        mcount, mamount, _, closed, new = latest_mig
        worsening, improving = [], []
        worsen_amt = improve_amt = 0.0
        for from_stage, to_stage in product(STAGES, STAGES):
            if from_stage == to_stage:
                continue
            n = mcount.loc[from_stage, to_stage]
            amt = mamount.loc[from_stage, to_stage]
            if n == 0:
                continue
            entry = f"{n} loan{'s' if n != 1 else ''} worth Rs. {amt:,.2f} Cr slipped from {from_stage} to {to_stage}"
            if STAGES.index(to_stage) > STAGES.index(from_stage):
                worsening.append(entry)
                worsen_amt += amt
            else:
                improving.append(
                    f"{n} loan{'s' if n != 1 else ''} worth Rs. {amt:,.2f} Cr recovered from {from_stage} to {to_stage}")
                improve_amt += amt

        lines.append("\nWHAT MOVED")
        if worsening:
            lines.append(f"Deterioration — Rs. {worsen_amt:,.2f} Cr in total:")
            lines.extend(f"  - {e}" for e in worsening)
        if improving:
            lines.append(f"Improvement — Rs. {improve_amt:,.2f} Cr in total:")
            lines.extend(f"  - {e}" for e in improving)
        if not worsening and not improving:
            lines.append("No loan changed stage between the two periods.")

        net = improve_amt - worsen_amt
        if worsening or improving:
            lines.append(
                f"On balance, {'improvement outweighed deterioration' if net > 0 else 'deterioration outweighed improvement'} "
                f"by Rs. {abs(net):,.2f} Cr."
            )

        wo_ids = set(writeoffs["LoanID"]) if writeoffs is not None and len(writeoffs) else set()
        closed_repaid = closed[~closed["LoanID"].isin(wo_ids)] if len(closed) else closed

        if len(closed_repaid) or len(new):
            lines.append("\nBOOK MOVEMENT")
            if len(closed_repaid):
                lines.append(f"{len(closed_repaid)} loan(s) left the book through repayment or closure, "
                              f"Rs. {closed_repaid['Outstanding_Cr'].sum():,.2f} Cr as last recorded.")
            if len(new):
                lines.append(f"{len(new)} new loan(s) were originated, "
                              f"Rs. {new['Outstanding_Cr'].sum():,.2f} Cr.")

    # ---- 4. Write-offs ----
    if writeoffs is not None and len(writeoffs):
        lines.append("\nWRITE-OFFS")
        lines.append(
            f"{len(writeoffs)} loan(s) were written off in {period}, reducing the loss allowance by "
            f"Rs. {writeoffs['WriteOff_Amount_Cr'].sum():,.3f} Cr. These have been removed from the "
            f"active portfolio and are shown separately here rather than under repayments, since a "
            f"write-off carries no cash recovery."
        )

    # ---- 4b. By product ----
    if product_summary is not None and not product_summary.empty:
        lines.append("\nBY PRODUCT")
        top_by_book = product_summary.sort_values("Outstanding_Cr_to", ascending=False).iloc[0]
        lines.append(
            f"{top_by_book['Product']} is the largest product line at {period}, with "
            f"Rs. {top_by_book['Outstanding_Cr_to']:,.2f} Cr outstanding."
        )
        worsened = product_summary[product_summary["Δ Stage3 Share (pp)"] > 0.5].sort_values(
            "Δ Stage3 Share (pp)", ascending=False)
        improved = product_summary[product_summary["Δ Stage3 Share (pp)"] < -0.5].sort_values(
            "Δ Stage3 Share (pp)")
        if len(worsened):
            w = worsened.iloc[0]
            lines.append(
                f"{w['Product']} showed the sharpest rise in Stage 3 share, up "
                f"{w['Δ Stage3 Share (pp)']:.2f} pp to {w['Stage3_Share_pct_to']:.2f}% — "
                f"this product line is worth a closer look."
            )
        if len(improved):
            imp = improved.iloc[0]
            lines.append(
                f"{imp['Product']} improved the most, with Stage 3 share down "
                f"{abs(imp['Δ Stage3 Share (pp)']):.2f} pp to {imp['Stage3_Share_pct_to']:.2f}%."
            )
        if not len(worsened) and not len(improved):
            lines.append("Stage 3 share was broadly stable across all product lines.")
        lines.append("Full product-wise figures are on the Product summary sheet; branch-wise figures "
                     "are on the Branch sheet, which can be filtered to any branch.")

    # ---- 5. What the Controller should look at ----
    watch = []
    if prior is not None:
        s3_delta = latest["Stage 3_Share_pct"] - prior["Stage 3_Share_pct"]
        s2_delta = latest["Stage 2_Share_pct"] - prior["Stage 2_Share_pct"]
        if s3_delta > 1.0:
            watch.append(f"Stage 3 share rose {s3_delta:.2f} pp — review recovery and provisioning adequacy.")
        if s2_delta > 1.0:
            watch.append(f"Stage 2 share rose {s2_delta:.2f} pp — early stress building; review the SICR triggers behind it.")
        if latest["Overall_PCR_pct"] < prior["Overall_PCR_pct"] - 1.0:
            watch.append("Coverage fell more than 1 pp — confirm this reflects genuine recovery, not under-provisioning.")
    if watch:
        lines.append("\nPOINTS FOR ATTENTION")
        lines.extend(f"  - {w}" for w in watch)

    lines.append(
        "\nBasis: figures are taken as computed under Ind AS 109 and RBI norms in the source data. "
        "This draft has not independently verified them and requires Controller review before "
        "Board or Audit Committee submission."
    )
    return "\n".join(lines)


# --------------------------------------------------------------------------
# STEP 5 (optional): AI-polished narrative - aggregated figures only
# --------------------------------------------------------------------------
def generate_ai_commentary(summary: pd.DataFrame, rule_based_draft: str) -> str:
    try:
        import anthropic
    except ImportError:
        return "[AI mode skipped: run 'pip install anthropic']\n\n" + rule_based_draft

    api_key = os.environ.get("ANTHROPIC_API_KEY")
    if not api_key:
        return "[AI mode skipped: ANTHROPIC_API_KEY not set]\n\n" + rule_based_draft

    latest = summary.iloc[-1].to_dict()
    prior = summary.iloc[-2].to_dict() if len(summary) > 1 else None

    prompt = f"""You are assisting a Financial Controller at an Indian NBFC drafting ECL
provisioning commentary for the Audit Committee/Board under Ind AS 109.
Use ONLY the aggregated figures below - no customer names, no invented numbers.

Current period: {latest}
Prior period: {prior}

Write a factual 150-200 word Board-ready paragraph covering overall coverage,
stage-wise movement, and any deterioration signal worth flagging."""

    client = anthropic.Anthropic(api_key=api_key)
    response = client.messages.create(
        model="claude-sonnet-4-6", max_tokens=500,
        messages=[{"role": "user", "content": prompt}],
    )
    ai_text = "".join(b.text for b in response.content if hasattr(b, "text"))
    return ("ECL PROVISIONING COMMENTARY (AI-DRAFTED — REQUIRES CONTROLLER REVIEW)\n"
            + "=" * 70 + "\n" + ai_text.strip() + "\n" + "=" * 70)


# --------------------------------------------------------------------------
# STEP 6: Write Excel workbook with clickable drill-down
# --------------------------------------------------------------------------
def safe_cell(value):
    """
    Excel treats a text cell beginning with =, +, - or @ as a formula, which
    corrupts the workbook and triggers the "we found a problem with some
    content" repair prompt on open. Any such text is prefixed with a single
    space so it is stored as plain text. Numbers and dates pass through
    untouched.
    """
    if isinstance(value, str) and value[:1] in ("=", "+", "-", "@"):
        return " " + value
    return value


def safe_sheet_name(name: str) -> str:
    return name[:31]


def write_product_summary_sheet(wb, period_from: str, period_to: str, summary_df: pd.DataFrame):
    """
    Compact Product-wise table: few rows (a handful of products), so this is
    shown in full - no filter needed. Outstanding, provision, loan count and
    Stage 3 share for both periods, with deltas, sorted by current outstanding.
    """
    ws = wb.create_sheet(safe_sheet_name(f"Product_{period_from}_{period_to}"))
    ws["A1"] = f"Product-wise Summary: {period_from} {ARROW} {period_to}"
    ws["A1"].font = TITLE_FONT
    ws["A2"] = "Every product line is shown in full below (no filter needed - see the Branch sheet for that)."
    ws["A2"].font = SUBTITLE_FONT

    cols = ["Product", f"Loans_from", f"Loans_to",
            "Outstanding_Cr_from", "Outstanding_Cr_to", "Δ Outstanding_Cr",
            "Provision_Cr_from", "Provision_Cr_to", "Δ Provision_Cr",
            "Stage3_Share_pct_from", "Stage3_Share_pct_to", "Δ Stage3 Share (pp)"]
    headers = ["Product", f"Loans ({period_from})", f"Loans ({period_to})",
               f"Outstanding ({period_from})", f"Outstanding ({period_to})", "Δ Outstanding",
               f"Provision ({period_from})", f"Provision ({period_to})", "Δ Provision",
               f"Stage 3 Share % ({period_from})", f"Stage 3 Share % ({period_to})", "Δ Stage 3 Share (pp)"]

    hr = 4
    for j, h in enumerate(headers, start=1):
        c = ws.cell(row=hr, column=j, value=h)
        c.font, c.fill = HEADER_FONT, HEADER_FILL
        c.alignment = Alignment(horizontal="center", wrap_text=True)

    for i, (_, row) in enumerate(summary_df.iterrows(), start=hr + 1):
        for j, col in enumerate(cols, start=1):
            val = row[col]
            header = headers[j - 1]
            if col == "Δ Outstanding_Cr":
                val = f"=E{i}-D{i}"          # Outstanding_to (E) - Outstanding_from (D)
            elif col == "Δ Provision_Cr":
                val = f"=H{i}-G{i}"          # Provision_to (H) - Provision_from (G)
            elif col == "Δ Stage3 Share (pp)":
                val = f"=K{i}-J{i}"          # Stage3_Share_to (K) - Stage3_Share_from (J)
            cell = ws.cell(row=i, column=j, value=safe_cell(val) if not isinstance(val, str) or not val.startswith("=") else val)
            cell.border = THIN_BORDER
            if col == "Δ Outstanding_Cr":
                cell.number_format = "#,##0.00"
            elif col == "Δ Provision_Cr":
                cell.number_format = "#,##0.000"
            elif col == "Δ Stage3 Share (pp)":
                cell.number_format = "0.00"
            if "Δ" in header and "Stage 3" in header:
                orig_val = row[col]
                if isinstance(orig_val, (int, float)) and orig_val > 0:
                    cell.font = Font(color="FFC00000", bold=True)
    for j in range(1, len(headers) + 1):
        ws.column_dimensions[get_column_letter(j)].width = 18
    ws.column_dimensions["A"].width = 22
    ws.freeze_panes = ws.cell(row=hr + 1, column=2)
    ws.page_setup.orientation = "landscape"
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.sheet_properties.pageSetUpPr.fitToPage = True


def write_branch_filter_sheet(wb, period_from: str, period_to: str, df: pd.DataFrame):
    """
    Branch-wise detail as a genuine Excel TABLE (ListObject), not a plain
    range with hand-written formulas. This matters for one reason
    specifically: a plain range's SUBTOTAL formulas are anchored to whatever
    row count existed when the file was generated - if someone later pastes
    or types more rows directly into the sheet in Excel, those formulas do
    NOT expand to include them.

    An Excel Table behaves differently by design: it auto-expands its range
    the moment a new row is added immediately below it (typing into the row
    right under the table, or pasting rows there, pulls them into the
    table automatically), and its Total Row - which we turn on below, using
    SUM/COUNT functions that Excel implements internally via SUBTOTAL so
    they already ignore filtered-out rows - moves down and keeps working
    without any manual edit to a formula or a range reference.

    (Every run of this script still regenerates the sheet fresh from the
    current data, so this only matters if the person edits the workbook by
    hand afterwards - but that is exactly the scenario being guarded against.)
    """
    if "Branch" not in df.columns:
        return

    extra_cols = [c for c in ("Product",) if c in df.columns]

    df_from = df[df["Period"] == period_from][
        ["LoanID", "CustomerName", "Branch", "Stage", "Outstanding_Cr", "Provision_Cr"] + extra_cols
    ]
    df_to = df[df["Period"] == period_to][["LoanID", "Stage", "Outstanding_Cr", "Provision_Cr"]]
    merged = df_from.merge(df_to, on="LoanID", suffixes=("_from", "_to"), how="outer")

    merged["Branch"] = merged["Branch"].fillna("(unknown)")
    for c in ("CustomerName",) + tuple(extra_cols):
        if c in merged.columns:
            merged[c] = merged[c].fillna("")
    merged["Stage_from"] = merged["Stage_from"].fillna("(new loan)")
    merged["Stage_to"] = merged["Stage_to"].fillna("(exited)")
    merged["Outstanding_Cr_from"] = merged["Outstanding_Cr_from"].fillna(0)
    merged["Outstanding_Cr_to"] = merged["Outstanding_Cr_to"].fillna(0)
    merged["Provision_Cr_from"] = merged["Provision_Cr_from"].fillna(0)
    merged["Provision_Cr_to"] = merged["Provision_Cr_to"].fillna(0)

    headers = ["LoanID", "CustomerName", "Branch"] + extra_cols + [
        f"Stage ({period_from})", f"Stage ({period_to})",
        f"Outstanding ({period_from})", f"Outstanding ({period_to})",
        f"Provision ({period_from})", f"Provision ({period_to})",
    ]
    data_cols = ["LoanID", "CustomerName", "Branch"] + extra_cols + [
        "Stage_from", "Stage_to", "Outstanding_Cr_from", "Outstanding_Cr_to",
        "Provision_Cr_from", "Provision_Cr_to",
    ]
    numeric_cols = {"Outstanding_Cr_from", "Outstanding_Cr_to", "Provision_Cr_from", "Provision_Cr_to"}

    ws = wb.create_sheet(safe_sheet_name(f"Branch_{period_from}_{period_to}"))
    ws["A1"] = f"Branch-wise Detail (filterable): {period_from} {ARROW} {period_to}"
    ws["A1"].font = TITLE_FONT
    ws["A2"] = ("This is a live Excel Table. Use the filter dropdown on Branch (or Product/Stage) to "
               "isolate any subset - the Total row updates automatically to the visible rows, and the "
               "table itself grows automatically if more rows are added directly beneath it.")
    ws["A2"].font = SUBTITLE_FONT

    hr = 4
    first_data_row = hr + 1
    last_data_row = hr + len(merged)

    for j, h in enumerate(headers, start=1):
        c = ws.cell(row=hr, column=j, value=h)
        c.font, c.fill = HEADER_FONT, HEADER_FILL
        c.alignment = Alignment(horizontal="center", wrap_text=True)

    for i, (_, row) in enumerate(merged.iterrows(), start=first_data_row):
        for j, col in enumerate(data_cols, start=1):
            val = row[col]
            cell = ws.cell(row=i, column=j, value=safe_cell(val))
            if col in ("Outstanding_Cr_from", "Outstanding_Cr_to"):
                cell.number_format = "#,##0.00"
            elif col in ("Provision_Cr_from", "Provision_Cr_to"):
                cell.number_format = "#,##0.000"

    last_col_letter = get_column_letter(len(headers))
    total_display_row = last_data_row + 1
    table_name = re.sub(r"\W", "_", f"BranchDetail_{period_from}_{period_to}")[:255]
    tab = Table(displayName=table_name, ref=f"A{hr}:{last_col_letter}{total_display_row}")
    tab.tableStyleInfo = TableStyleInfo(name="TableStyleMedium9", showRowStripes=True,
                                        showFirstColumn=False, showLastColumn=False,
                                        showColumnStripes=False)

    # openpyxl does not auto-populate tableColumns from the sheet's header
    # cells - it must be done explicitly, with names matching exactly what
    # is already written in the header row, or Excel will reject the file.
    tab._initialise_columns()
    for tcol, name in zip(tab.tableColumns, headers):
        tcol.name = name

    # _initialise_columns() also defaults the AutoFilter range to the WHOLE
    # table ref, including the totals row - per the OOXML spec the filter
    # range should stop at the last DATA row, excluding totals. Left as the
    # openpyxl default, this is a latent source of exactly the "we found a
    # problem with some content" repair prompt seen earlier in this project,
    # so it is set explicitly here instead.
    tab.autoFilter = AutoFilter(ref=f"A{hr}:{last_col_letter}{last_data_row}")

    # Turn on the native Total Row: Excel auto-extends this WITH the table
    # as rows are added, and its aggregate functions already respect any
    # active filter - no hand-written formula or fixed range involved.
    tab.totalsRowShown = True
    tab.totalsRowCount = 1
    for idx, col in enumerate(data_cols):
        tcol = tab.tableColumns[idx]
        if idx == 0:
            tcol.totalsRowFunction = "count"
        elif col in numeric_cols:
            tcol.totalsRowFunction = "sum"
        else:
            tcol.totalsRowFunction = "none"
    ws.add_table(tab)

    # IMPORTANT: setting totalsRowFunction on the Table's metadata alone does
    # NOT make Excel display a value - openpyxl does not write the
    # corresponding formula into the sheet, so the row renders BLANK unless
    # the formula is put there explicitly. Excel's own "Total Row" feature
    # writes a SUBTOTAL formula (which already respects any active filter,
    # ignoring hidden/filtered-out rows) when a user turns it on via the UI;
    # replicating that here so the row actually shows something the moment
    # the file is opened, in any application, not only after a user
    # manually toggles the Total Row checkbox in Excel.
    ws.cell(row=total_display_row, column=1, value="Total").font = Font(bold=True)
    for idx, col in enumerate(data_cols, start=1):
        if idx == 1:
            continue
        col_letter = get_column_letter(idx)
        if col in numeric_cols:
            formula = f"=SUBTOTAL(109,{col_letter}{first_data_row}:{col_letter}{last_data_row})"
            cell = ws.cell(row=total_display_row, column=idx, value=formula)
            cell.font = Font(bold=True)
            cell.number_format = "#,##0.00" if "Outstanding" in col else "#,##0.000"

    ws.freeze_panes = ws.cell(row=first_data_row, column=2)

    # ---- Print setup: this sheet is wide (loan detail + two periods'
    # worth of stage/outstanding/provision columns). Left at default
    # portrait/100% scale, columns run off the printable page width and
    # split a single logical row across two disconnected pages, losing the
    # LoanID/CustomerName on the continuation page - landscape + fit-to-
    # width keeps every column, including the identifying ones, on one
    # printed page width (rows still paginate normally, which is fine).
    ws.page_setup.orientation = "landscape"
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.print_title_cols = "A:D" if extra_cols else "A:C"  # repeat ID columns on every printed page

    for j, h in enumerate(headers, start=1):
        ws.column_dimensions[get_column_letter(j)].width = max(11, min(len(h) + 2, 16))
    ws.column_dimensions["B"].width = 20      # CustomerName
    ws.column_dimensions["C"].width = 14      # Branch
    if extra_cols:
        ws.column_dimensions["D"].width = 18  # Product - "Equipment Finance" etc. need the room



def write_iracp_loan_working_sheet(wb, period: str, loan_detail: pd.DataFrame):
    """
    Writes the per-loan IRACP audit trail with the provision computed by a
    LIVE EXCEL FORMULA, not a value baked in by Python:

        Unsecured_Cr    = Outstanding_Cr - Secured_Cr                [formula]
        IRACP Provision = Secured_Cr * Rate_Secured + Unsecured_Cr * Rate_Unsecured  [formula]

    This single pattern covers every asset category (see
    compute_iracp_provision_for_loan for why) - Standard/Sub-standard/Loss
    simply carry their whole outstanding in the "unsecured" column at their
    flat rate, with secured = 0, so the same formula still gives the right
    answer. The "Basis" column states in plain words which rule fired for
    that row, and "Rate" columns are shown as actual percentages.

    Returns (sheet_name, cols) where cols maps field name -> column letter,
    so the calling category-level sheet can build SUMIF formulas against
    this sheet's live-computed Provision column instead of a Python total.
    """
    headers = ["LoanID", "CustomerName", "Ind AS 109 Stage", "DPD", "IRACP_Category",
               "Outstanding_Cr", "Secured_Cr", "Unsecured_Cr",
               "Rate_Secured_pct", "Rate_Unsecured_pct",
               "IRACP Provision (Rs. Cr)", "Ind AS 109 ECL (Rs. Cr)",
               "Security_Basis", "Basis (rule applied)"]
    col = {h: get_column_letter(i + 1) for i, h in enumerate(headers)}

    sheet_name = safe_sheet_name(f"IRACP_LoanWorking_{period}")
    ws = wb.create_sheet(sheet_name)
    ws["A1"] = f"IRACP Loan-level Working — {period}"
    ws["A1"].font = TITLE_FONT
    ws["A2"] = ("Unsecured_Cr and the Provision column are live formulas (click any cell to see them) - "
               "change Outstanding_Cr or Secured_Cr and the provision recalculates automatically. "
               "Standard/Sub-standard/Loss rows carry their full outstanding under 'Unsecured_Cr' at "
               "their flat rate (Secured_Cr = 0) so one formula pattern works for every category.")
    ws["A2"].font = SUBTITLE_FONT

    hr = 4
    for j, h in enumerate(headers, start=1):
        c = ws.cell(row=hr, column=j, value=h)
        c.font, c.fill = HEADER_FONT, HEADER_FILL
        c.alignment = Alignment(horizontal="center", wrap_text=True)

    first_row = hr + 1
    for i, (_, row) in enumerate(loan_detail.iterrows(), start=first_row):
        ws.cell(row=i, column=1, value=safe_cell(row["LoanID"]))
        ws.cell(row=i, column=2, value=safe_cell(row["CustomerName"]))
        ws.cell(row=i, column=3, value=safe_cell(row["Ind AS 109 Stage"]))
        ws.cell(row=i, column=4, value=row["DPD"])
        ws.cell(row=i, column=5, value=safe_cell(row["IRACP_Category"]))

        c_out = ws.cell(row=i, column=6, value=round(float(row["Outstanding_Cr"]), 4))
        c_out.number_format = "#,##0.0000"
        c_sec = ws.cell(row=i, column=7, value=round(float(row["Secured_Cr"]), 4))
        c_sec.number_format = "#,##0.0000"

        # Unsecured_Cr - LIVE FORMULA: Outstanding minus Secured
        c_unsec = ws.cell(row=i, column=8, value=f"={col['Outstanding_Cr']}{i}-{col['Secured_Cr']}{i}")
        c_unsec.number_format = "#,##0.0000"

        c_rs = ws.cell(row=i, column=9, value=float(row["Rate_Secured_pct"]))
        c_rs.number_format = "0.00%"
        c_ru = ws.cell(row=i, column=10, value=float(row["Rate_Unsecured_pct"]))
        c_ru.number_format = "0.00%"

        # IRACP Provision - LIVE FORMULA
        formula = (f"={col['Secured_Cr']}{i}*{col['Rate_Secured_pct']}{i}"
                   f"+{col['Unsecured_Cr']}{i}*{col['Rate_Unsecured_pct']}{i}")
        c_prov = ws.cell(row=i, column=11, value=formula)
        c_prov.number_format = "#,##0.0000"

        c_ecl = ws.cell(row=i, column=12, value=round(float(row["Ind AS 109 ECL (Rs. Cr)"]), 4))
        c_ecl.number_format = "#,##0.0000"
        ws.cell(row=i, column=13, value=safe_cell(row["Security_Basis"]))
        ws.cell(row=i, column=14, value=safe_cell(row["Basis"]))

        for cc in range(1, len(headers) + 1):
            ws.cell(row=i, column=cc).border = THIN_BORDER

    last_row = first_row + len(loan_detail) - 1
    ws.auto_filter.ref = f"A{hr}:{get_column_letter(len(headers))}{max(last_row, hr)}"
    ws.freeze_panes = ws.cell(row=first_row, column=2)
    for j in range(1, len(headers) + 1):
        ws.column_dimensions[get_column_letter(j)].width = 16
    ws.column_dimensions["B"].width = 20
    ws.column_dimensions["N"].width = 40

    ws.page_setup.orientation = "landscape"
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.print_title_cols = "A:B"

    return sheet_name, col, first_row, last_row


def write_dataframe_sheet(wb, sheet_name, title, subtitle, df, start_row=4, index_label=None,
                          formula_total_col=None):
    """
    Generic helper: writes a titled DataFrame (with its index if named) to a
    new sheet. If formula_total_col names a column, that column is written
    as a live Excel SUM formula across the OTHER numeric columns in the same
    row, instead of the value pandas already computed - so opening the cell
    shows the actual arithmetic, not just its result.
    """
    ws = wb.create_sheet(safe_sheet_name(sheet_name))
    ws["A1"] = title
    ws["A1"].font = TITLE_FONT
    if subtitle:
        ws["A2"] = subtitle
        ws["A2"].font = SUBTITLE_FONT

    df_to_write = df.reset_index() if index_label else df
    if index_label:
        df_to_write = df_to_write.rename(columns={"index": index_label})

    columns = list(df_to_write.columns)
    for j, col in enumerate(columns, start=1):
        c = ws.cell(row=start_row, column=j, value=str(col))
        c.font = HEADER_FONT
        c.fill = HEADER_FILL
        c.alignment = Alignment(horizontal="center")

    total_col_idx = columns.index(formula_total_col) + 1 if formula_total_col in columns else None
    numeric_cols_idx = [j for j, col in enumerate(columns, start=1)
                        if col != formula_total_col and col != index_label
                        and pd.api.types.is_numeric_dtype(df_to_write[col])] if total_col_idx else []

    for i, (_, row) in enumerate(df_to_write.iterrows(), start=start_row + 1):
        for j, col in enumerate(columns, start=1):
            if j == total_col_idx and numeric_cols_idx:
                first_letter = get_column_letter(numeric_cols_idx[0])
                last_letter = get_column_letter(numeric_cols_idx[-1])
                cell = ws.cell(row=i, column=j, value=f"=SUM({first_letter}{i}:{last_letter}{i})")
            else:
                val = row[col]
                cell = ws.cell(row=i, column=j, value=safe_cell(val))
            cell.border = THIN_BORDER
            if isinstance(row[col], str) and (row[col].startswith("TOTAL") or row[col] == "Closing balance"):
                cell.font = Font(bold=True)
    for j, col in enumerate(columns, start=1):
        ws.column_dimensions[get_column_letter(j)].width = max(20, len(str(col)) + 2)
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    return ws


def write_excel(summary: pd.DataFrame, commentary: str, transitions: list, out_path: str,
                 indas107_tables: list = None, iracp_table=None, iracp_period=None,
                 iracp_method_note: str = None, iracp_assumptions=None, iracp_loan_detail=None,
                 dpd_migrations: dict = None, writeoffs_by_period: dict = None,
                 df_active: pd.DataFrame = None):
    wb = Workbook()

    # ---- Summary sheet ----
    ws = wb.active
    ws.title = "Summary"
    ws["A1"] = "ECL Provisioning Summary & Commentary"
    ws["A1"].font = TITLE_FONT
    ws["A2"] = f"Generated: {datetime.now().strftime('%d-%b-%Y %H:%M')}"
    ws["A2"].font = SUBTITLE_FONT

    # ------------------------------------------------------------------
    # COMPACT ONE-PAGE LAYOUT
    # Periods run ACROSS as columns (usually just two), metrics run DOWN as
    # rows. This keeps the whole summary inside a single printable page
    # instead of sprawling out to column Q.
    # ------------------------------------------------------------------
    periods_list = list(summary["Period"])
    S = {row["Period"]: row for _, row in summary.iterrows()}

    metric_rows = [
        ("PORTFOLIO", None, None),
        ("Number of loans", "Loan_Count", "#,##0"),
        ("Outstanding (Rs. Cr)", "Total_Outstanding_Cr", "#,##0.00"),
        ("ECL provision held (Rs. Cr)", "Total_Provision_Cr", "#,##0.000"),
        ("Provision coverage (%)", "Overall_PCR_pct", "0.00"),
        ("STAGE 1", None, None),
        ("  Loans", "Stage 1_Count", "#,##0"),
        ("  Outstanding (Rs. Cr)", "Stage 1_Outstanding_Cr", "#,##0.00"),
        ("  Provision (Rs. Cr)", "Stage 1_Provision_Cr", "#,##0.000"),
        ("  Share of book (%)", "Stage 1_Share_pct", "0.00"),
        ("STAGE 2", None, None),
        ("  Loans", "Stage 2_Count", "#,##0"),
        ("  Outstanding (Rs. Cr)", "Stage 2_Outstanding_Cr", "#,##0.00"),
        ("  Provision (Rs. Cr)", "Stage 2_Provision_Cr", "#,##0.000"),
        ("  Share of book (%)", "Stage 2_Share_pct", "0.00"),
        ("STAGE 3", None, None),
        ("  Loans", "Stage 3_Count", "#,##0"),
        ("  Outstanding (Rs. Cr)", "Stage 3_Outstanding_Cr", "#,##0.00"),
        ("  Provision (Rs. Cr)", "Stage 3_Provision_Cr", "#,##0.000"),
        ("  Share of book (%)", "Stage 3_Share_pct", "0.00"),
    ]

    hr = 4
    c = ws.cell(row=hr, column=1, value="Metric")
    c.font, c.fill = HEADER_FONT, HEADER_FILL
    for j, per in enumerate(periods_list, start=2):
        c = ws.cell(row=hr, column=j, value=per)
        c.font, c.fill = HEADER_FONT, HEADER_FILL
        c.alignment = Alignment(horizontal="center")
    chg_col = len(periods_list) + 2
    if len(periods_list) == 2:
        c = ws.cell(row=hr, column=chg_col, value="Change")
        c.font, c.fill = HEADER_FONT, HEADER_FILL
        c.alignment = Alignment(horizontal="center")

    r = hr + 1
    for label, key, fmt in metric_rows:
        if key is None:                      # section band
            c = ws.cell(row=r, column=1, value=label)
            c.font = Font(bold=True, color="FF1F4E78")
            c.fill = PatternFill(start_color="FFEAF0F8", end_color="FFEAF0F8", fill_type="solid")
            for j in range(2, chg_col + 1):
                ws.cell(row=r, column=j).fill = PatternFill(
                    start_color="FFEAF0F8", end_color="FFEAF0F8", fill_type="solid")
            r += 1
            continue

        ws.cell(row=r, column=1, value=label).border = THIN_BORDER
        vals = []
        for j, per in enumerate(periods_list, start=2):
            v = S[per][key] if key in S[per] else None
            vals.append(v)
            cell = ws.cell(row=r, column=j, value=v)
            cell.border = THIN_BORDER
            cell.number_format = fmt
            cell.alignment = Alignment(horizontal="right")
        if len(periods_list) == 2 and all(v is not None for v in vals):
            # Change - live formula (later period column minus earlier),
            # not a value computed only in Python.
            from_letter = get_column_letter(2)
            to_letter = get_column_letter(3)
            delta = vals[1] - vals[0]  # only used to decide the highlight color below
            cell = ws.cell(row=r, column=chg_col, value=f"={to_letter}{r}-{from_letter}{r}")
            cell.border = THIN_BORDER
            cell.number_format = fmt
            cell.alignment = Alignment(horizontal="right")
            cell.font = Font(bold=True,
                             color="FFC00000" if delta > 0 and "Stage 3" in label else "FF000000")
        r += 1

    ws.column_dimensions["A"].width = 30
    for j in range(2, chg_col + 1):
        ws.column_dimensions[get_column_letter(j)].width = 15

    # ---- Write-offs summary line, if any ----
    if writeoffs_by_period:
        for per, wo in writeoffs_by_period.items():
            if wo is not None and len(wo):
                r += 1
                c = ws.cell(row=r, column=1, value=f"Write-offs in {per}")
                c.font = Font(bold=True, color="FFC00000")
                ws.cell(row=r, column=2,
                        value=f"{len(wo)} loan(s), Rs. {wo['WriteOff_Amount_Cr'].sum():,.3f} Cr")
                r += 1

    # ---- Commentary beneath the table ----
    r += 2
    ws.cell(row=r, column=1, value="COMMENTARY").font = Font(bold=True, size=12)
    r += 1
    for line in commentary.split("\n"):
        c = ws.cell(row=r, column=1, value=safe_cell(line))
        if line and not line.startswith((" ", "  ")) and line.isupper():
            c.font = Font(bold=True)
        r += 1

    # printable on one page
    ws.page_setup.orientation = "portrait"
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.freeze_panes = ws.cell(row=hr + 1, column=2)

    # ---- Per-transition: ONE migration sheet (stage + DPD) + ONE drill-down ----
    for (period_from, period_to, mcount, mamount, mprov, detail, closed, new) in transitions:
        mig_sheet_name = safe_sheet_name(f"Migration_{period_from}_{period_to}")
        drill_sheet_name = safe_sheet_name(f"Drilldown_{period_from}_{period_to}")

        ws_m = wb.create_sheet(mig_sheet_name)
        ws_m["A1"] = f"Migration Analysis: {period_from} \u2192 {period_to}"
        ws_m["A1"].font = TITLE_FONT
        ws_m["A2"] = ("Each table below is headed with the actual periods being compared. "
                      "Click any number to open the loan-level drill-down.")
        ws_m["A2"].font = SUBTITLE_FONT

        r = 4
        PERIOD_BAND_FILL = PatternFill(start_color="FFD6E4F0", end_color="FFD6E4F0", fill_type="solid")
        PERIOD_BAND_FONT = Font(bold=True, color="FF1F4E78")

        def _write_one(ws, top_row, caption, labels, matrix, worsen_fn,
                        numfmt=None, link=True):
            """
            Writes ONE From-x-To table starting at top_row; returns next free row.
            The periods themselves are baked into the header: a merged band
            above the column headers reads "AS AT <period_to>", and the
            corner cell reads "AS AT <period_from>" - so the periods being
            compared are visible on the table itself, not just in a
            separate sentence above it.
            """
            ws.cell(row=top_row, column=1, value=caption).font = Font(bold=True, italic=True)
            band_row = top_row + 1
            hr = band_row + 1
            tot_col = len(labels) + 2

            # merged band: "AS AT <period_to>" spanning the data + total columns
            ws.merge_cells(start_row=band_row, start_column=2, end_row=band_row, end_column=tot_col)
            band_cell = ws.cell(row=band_row, column=2, value=f"AS AT {period_to}")
            band_cell.font, band_cell.fill = PERIOD_BAND_FONT, PERIOD_BAND_FILL
            band_cell.alignment = Alignment(horizontal="center")
            corner_band = ws.cell(row=band_row, column=1, value=f"AS AT {period_from} \u25be")
            corner_band.font, corner_band.fill = PERIOD_BAND_FONT, PERIOD_BAND_FILL
            corner_band.alignment = Alignment(horizontal="center", wrap_text=True)

            c = ws.cell(row=hr, column=1, value="Stage/Bucket")
            c.font, c.fill = HEADER_FONT, HEADER_FILL
            for j, lbl in enumerate(labels, start=2):
                c = ws.cell(row=hr, column=j, value=lbl)
                c.font, c.fill = HEADER_FONT, HEADER_FILL
                c.alignment = Alignment(horizontal="center")
            c = ws.cell(row=hr, column=tot_col, value="Total")
            c.font, c.fill = HEADER_FONT, HEADER_FILL

            data_first_col = 2
            data_last_col = tot_col - 1
            for i, rl in enumerate(labels, start=hr + 1):
                c = ws.cell(row=i, column=1, value=rl)
                c.font, c.fill = HEADER_FONT, HEADER_FILL
                for j, cl in enumerate(labels, start=2):
                    v = matrix.loc[rl, cl]
                    v = int(v) if numfmt is None else round(float(v), 3)
                    cell = ws.cell(row=i, column=j, value=v)
                    cell.border = THIN_BORDER
                    cell.alignment = Alignment(horizontal="center")
                    if numfmt:
                        cell.number_format = numfmt
                    if rl == cl:
                        cell.fill = DIAG_FILL
                    elif worsen_fn(rl, cl):
                        cell.fill = WORSEN_FILL
                    else:
                        cell.fill = IMPROVE_FILL
                    if link and v:
                        cell.hyperlink = f"#'{drill_sheet_name}'!A1"
                        cell.font = LINK_FONT
                # ROW TOTAL - live formula summing this row's data cells
                row_letter_first = get_column_letter(data_first_col)
                row_letter_last = get_column_letter(data_last_col)
                tc = ws.cell(row=i, column=tot_col,
                             value=f"=SUM({row_letter_first}{i}:{row_letter_last}{i})")
                tc.font = Font(bold=True)
                tc.border = THIN_BORDER
                if numfmt:
                    tc.number_format = numfmt

            # COLUMN TOTALS - live formulas summing each column's data cells
            trow = hr + len(labels) + 1
            c = ws.cell(row=trow, column=1, value="Total")
            c.font, c.fill = HEADER_FONT, HEADER_FILL
            for j in range(data_first_col, tot_col + 1):
                col_letter = get_column_letter(j)
                cell = ws.cell(row=trow, column=j,
                               value=f"=SUM({col_letter}{hr + 1}:{col_letter}{trow - 1})")
                cell.font = Font(bold=True)
                cell.border = THIN_BORDER
                cell.alignment = Alignment(horizontal="center")
                if numfmt:
                    cell.number_format = numfmt
            return trow + 2

        def _write_block(ws, top_row, block_title, labels, m_prov, m_amt, m_cnt, worsen_fn):
            """
            Writes the three tables for one migration view, in the order:
                1. ECL provision (Rs. Cr)
                2. Outstanding   (Rs. Cr)
                3. Loan count
            """
            ws.cell(row=top_row, column=1, value=block_title).font = Font(bold=True, size=12)
            r = top_row + 1
            r = _write_one(ws, r, "1. ECL PROVISION HELD (Rs. Cr, as at later period)",
                           labels, m_prov, worsen_fn, numfmt="#,##0.000")
            r = _write_one(ws, r, "2. OUTSTANDING (Rs. Cr, as at later period)",
                           labels, m_amt, worsen_fn, numfmt="#,##0.00")
            r = _write_one(ws, r, "3. LOAN COUNT",
                           labels, m_cnt, worsen_fn, numfmt=None)
            return r

        # --- Block A: Ind AS 109 Stage migration ---
        r = _write_block(ws_m, r, "A. IND AS 109 STAGE MIGRATION",
                         STAGES, mprov, mamount, mcount,
                         lambda a, b: STAGES.index(b) > STAGES.index(a))

        # --- Block B: DPD ageing bucket migration ---
        dpd_mig = dpd_migrations.get((period_from, period_to)) if dpd_migrations else None
        if dpd_mig is not None:
            dcount, damount, dprov, _ = dpd_mig
            r = _write_block(ws_m, r, "B. DPD AGEING BUCKET MIGRATION",
                             DPD_BUCKET_NAMES, dprov, damount, dcount,
                             lambda a, b: DPD_BUCKET_NAMES.index(b) > DPD_BUCKET_NAMES.index(a))

        # --- Block 3: entries / exits (write-offs shown separately from
        #     ordinary repayment closures, to avoid double-listing the same
        #     loan under both) ---
        wo_this_period = writeoffs_by_period.get(period_to) if writeoffs_by_period else None
        wo_ids = set(wo_this_period["LoanID"]) if wo_this_period is not None and len(wo_this_period) else set()
        closed_repaid = closed[~closed["LoanID"].isin(wo_ids)] if len(closed) else closed

        ws_m.cell(row=r, column=1, value="C. PORTFOLIO ENTRIES & EXITS").font = Font(bold=True, size=12)
        ws_m.cell(row=r + 1, column=1,
                  value=f"Repaid / closed (non write-off) between {period_from} and {period_to}: "
                        f"{len(closed_repaid)} loan(s), Rs. {closed_repaid['Outstanding_Cr'].sum():.2f} Cr "
                        f"(as last recorded)")
        ws_m.cell(row=r + 2, column=1,
                  value=f"Newly originated between {period_from} and {period_to}: {len(new)} loan(s), "
                        f"Rs. {new['Outstanding_Cr'].sum():.2f} Cr")
        next_r = r + 3
        if wo_ids:
            ws_m.cell(row=next_r, column=1,
                      value=f"Written off in {period_to}: {len(wo_ids)} loan(s), "
                            f"Rs. {wo_this_period['WriteOff_Amount_Cr'].sum():.3f} Cr "
                            f"(see Ind AS 107 reconciliation — shown as a distinct movement, not a repayment)")
            next_r += 1
        r = next_r

        for col in "ABCDEFGHIJKLMN":
            ws_m.column_dimensions[col].width = 15

        # Print setup: Block B (DPD ageing) runs to 8 columns and, left at
        # default portrait/100%, loses its right-hand columns off the page
        # edge entirely when printed or exported to PDF (confirmed by
        # rendering this sheet before this fix was applied - the 91-120,
        # 120+ and Total columns of the provision matrix did not appear on
        # any page). Landscape + fit-to-width keeps every column, in every
        # block, on one printed page width.
        ws_m.page_setup.orientation = "landscape"
        ws_m.page_setup.fitToWidth = 1
        ws_m.page_setup.fitToHeight = 0
        ws_m.sheet_properties.pageSetUpPr.fitToPage = True
        ws_m.column_dimensions["A"].width = 18

        # --- ONE consolidated drill-down sheet (replaces 9 per-cell sheets) ---
        stage_drill = build_consolidated_drilldown(detail, period_from, period_to)
        ws_d = wb.create_sheet(drill_sheet_name)
        ws_d["A1"] = f"Loan-level Drill-down: {period_from} \u2192 {period_to}"
        ws_d["A1"].font = TITLE_FONT
        ws_d["A2"] = f"<< Back to {mig_sheet_name}"
        ws_d["A2"].hyperlink = f"#'{mig_sheet_name}'!A1"
        ws_d["A2"].font = LINK_FONT
        ws_d["A3"] = ("Every loan present in both periods, with its stage and DPD bucket in each. "
                      "Use Excel's filter on the Movement / Stage / DPD Bucket columns to isolate any cell "
                      "of the matrices.")
        ws_d["A3"].font = SUBTITLE_FONT

        # merge DPD bucket columns into the stage drill-down where available
        if dpd_mig is not None and not stage_drill.empty:
            _, _, _, dpd_flat = dpd_mig
            keep = ["LoanID",
                    f"DPD Bucket ({period_from})", f"DPD Bucket ({period_to})",
                    f"DPD ({period_from})", f"DPD ({period_to})"]
            stage_drill = stage_drill.merge(dpd_flat[keep], on="LoanID", how="left")

        if stage_drill.empty:
            ws_d["A5"] = "No loans common to both periods."
        else:
            hrow = 5
            for j, col in enumerate(stage_drill.columns, start=1):
                c = ws_d.cell(row=hrow, column=j, value=str(col))
                c.font, c.fill = HEADER_FONT, HEADER_FILL
            for i, (_, row) in enumerate(stage_drill.iterrows(), start=hrow + 1):
                for j, col in enumerate(stage_drill.columns, start=1):
                    cell = ws_d.cell(row=i, column=j, value=safe_cell(row[col]))
                    cell.border = THIN_BORDER
                    if col == "Movement":
                        if row[col] == "Deteriorated":
                            cell.fill = WORSEN_FILL
                        elif row[col] == "Improved":
                            cell.fill = IMPROVE_FILL
            ws_d.auto_filter.ref = (f"A{hrow}:"
                                     f"{get_column_letter(len(stage_drill.columns))}{hrow + len(stage_drill)}")
            ws_d.freeze_panes = ws_d.cell(row=hrow + 1, column=1)
            for j, col in enumerate(stage_drill.columns, start=1):
                ws_d.column_dimensions[get_column_letter(j)].width = max(15, min(len(str(col)) + 2, 28))
            # Override two columns known to hold values longer than their
            # header text (a width-from-header-length approach truncates
            # "Equipment Finance" under a "Product" header otherwise).
            for j, col in enumerate(stage_drill.columns, start=1):
                if col == "CustomerName":
                    ws_d.column_dimensions[get_column_letter(j)].width = 20
                elif col == "Product":
                    ws_d.column_dimensions[get_column_letter(j)].width = 19

            # Print setup: same page-width problem as the Migration sheet -
            # this table commonly runs to 9+ columns (Stage x2, Movement,
            # LoanID, CustomerName, Product, Branch, Outstanding x2,
            # Provision x2, DPD columns), which spilled LoanID/CustomerName
            # onto a disconnected page in testing. Landscape + fit-to-width,
            # with the identifying columns repeated on every printed page.
            ws_d.page_setup.orientation = "landscape"
            ws_d.page_setup.fitToWidth = 1
            ws_d.page_setup.fitToHeight = 0
            ws_d.sheet_properties.pageSetUpPr.fitToPage = True
            ws_d.print_title_cols = "A:E"

        # ---- Product-wise summary (compact - few products, shown in full) ----
        if df_active is not None and "Product" in df_active.columns:
            prod_summary = build_group_summary(df_active, period_from, period_to, "Product")
            if not prod_summary.empty:
                write_product_summary_sheet(wb, period_from, period_to, prod_summary)

        # ---- Branch-wise detail (filterable - branches may run into the
        #      hundreds, so a live-filter table beats one giant static pivot) ----
        if df_active is not None and "Branch" in df_active.columns:
            write_branch_filter_sheet(wb, period_from, period_to, df_active)

    # ---- STATUTORY DISCLOSURE: Ind AS 107 para 35H reconciliation ----
    if indas107_tables:
        for period_from, period_to, recon_df in indas107_tables:
            write_dataframe_sheet(
                wb,
                sheet_name=f"IndAS107_35H_{period_from}_{period_to}",
                title=f"Ind AS 107 para 35H — Loss Allowance Reconciliation: {period_from} \u2192 {period_to}",
                subtitle=("Movement in ECL by category. Mapping used: 12-month ECL = Stage 1, "
                          "Lifetime ECL (not credit-impaired) = Stage 2, Lifetime ECL (credit-impaired) "
                          "= Stage 3 — verify this mapping against your Board-approved SICR policy."),
                df=recon_df,
                index_label="Movement",
                formula_total_col="Total",
            )

    # ---- STATUTORY DISCLOSURE: RBI IRACP vs Ind AS 109 comparison ----
    if iracp_table is not None:
        # Write the loan-level working sheet FIRST so the category-level
        # table below can reference its live formulas via SUMIF, rather than
        # repeating a Python-computed total - keeping the whole chain,
        # loan -> category -> comparison, auditable by clicking cells.
        work_sheet_name, work_cols, work_first, work_last = (None, None, None, None)
        if iracp_loan_detail is not None:
            work_sheet_name, work_cols, work_first, work_last = write_iracp_loan_working_sheet(
                wb, iracp_period, iracp_loan_detail)

        ws_i = wb.create_sheet(safe_sheet_name(f"RBI_IRACP_Comparison_{iracp_period}"))
        ws_i["A1"] = f"RBI IRACP vs Ind AS 109 Provisioning Comparison — {iracp_period}"
        ws_i["A1"].font = TITLE_FONT
        ws_i["A2"] = ((iracp_method_note or "") + " Per RBI (NBFC – IRACP) Directions, 2025: where Ind AS 109 "
                     "ECL is lower than the IRACP floor, the shortfall is appropriated from net profit "
                     "after tax to a separate 'Impairment Reserve' (not reckoned for regulatory capital)."
                     + (f" Category totals below are live SUMIF formulas against the "
                        f"'{work_sheet_name}' sheet." if work_sheet_name else ""))
        ws_i["A2"].font = SUBTITLE_FONT

        headers_i = list(iracp_table.columns)
        hr_i = 4
        for j, h in enumerate(headers_i, start=1):
            c = ws_i.cell(row=hr_i, column=j, value=h)
            c.font, c.fill = HEADER_FONT, HEADER_FILL
            c.alignment = Alignment(horizontal="center", wrap_text=True)

        numeric_formula_cols = {"Loan_Count", "Outstanding_Cr", "Secured_Cr", "Unsecured_Cr",
                                 "Ind_AS_109_ECL_Provision_Cr", "IRACP_Provision_Cr"}
        col_num_i = {h: j for j, h in enumerate(headers_i, start=1)}
        col_letter_i = {h: get_column_letter(j) for j, h in enumerate(headers_i, start=1)}
        # map each comparison-table numeric column to its counterpart column
        # on the loan-working sheet, for the SUMIF criteria range/sum range
        source_col_map = {
            "Loan_Count": "LoanID", "Outstanding_Cr": "Outstanding_Cr",
            "Secured_Cr": "Secured_Cr", "Unsecured_Cr": "Unsecured_Cr",
            "Ind_AS_109_ECL_Provision_Cr": "Ind AS 109 ECL (Rs. Cr)",
            "IRACP_Provision_Cr": "IRACP Provision (Rs. Cr)",
        }

        for i, (_, row) in enumerate(iracp_table.iterrows(), start=hr_i + 1):
            cat = row["IRACP Asset Category"]
            for h in headers_i:
                cell = ws_i.cell(row=i, column=col_num_i[h])
                if h == "IRACP Asset Category" or h == "Rate(s) Applied":
                    cell.value = safe_cell(row[h])
                    continue
                if h == "Excess/(Shortfall) of Ind AS 109 over IRACP_Cr":
                    if work_sheet_name and cat != "TOTAL":
                        cell.value = (f"={col_letter_i['Ind_AS_109_ECL_Provision_Cr']}{i}"
                                     f"-{col_letter_i['IRACP_Provision_Cr']}{i}")
                    elif work_sheet_name and cat == "TOTAL":
                        cell.value = f"=SUM({col_letter_i[h]}{hr_i + 1}:{col_letter_i[h]}{i - 1})"
                    else:
                        cell.value = round(float(row[h]), 3)
                    cell.number_format = "#,##0.000"
                    continue
                if h in numeric_formula_cols and work_sheet_name:
                    src_col = source_col_map[h]
                    src_letter = work_cols[src_col]
                    cat_letter = work_cols["IRACP_Category"]
                    if cat == "TOTAL":
                        # total row: sum of the category rows above it, so it
                        # stays correct even if a category has zero loans
                        # (and therefore no row) in a given run
                        cell.value = f"=SUM({col_letter_i[h]}{hr_i + 1}:{col_letter_i[h]}{i - 1})"
                    else:
                        func = "COUNTIF" if h == "Loan_Count" else "SUMIF"
                        if func == "COUNTIF":
                            cell.value = (f"=COUNTIF({work_sheet_name}!{cat_letter}{work_first}:"
                                         f"{cat_letter}{work_last},A{i})")
                        else:
                            cell.value = (f"=SUMIF({work_sheet_name}!{cat_letter}{work_first}:"
                                         f"{cat_letter}{work_last},A{i},{work_sheet_name}!{src_letter}"
                                         f"{work_first}:{src_letter}{work_last})")
                    cell.number_format = "#,##0" if h == "Loan_Count" else "#,##0.000"
                else:
                    cell.value = round(float(row[h]), 3) if isinstance(row[h], (int, float)) else safe_cell(row[h])
                cell.border = THIN_BORDER

        for j in range(1, len(headers_i) + 1):
            ws_i.column_dimensions[get_column_letter(j)].width = 20
        ws_i.column_dimensions[col_letter_i.get("Rate(s) Applied", "B")].width = 45
        ws_i.page_setup.orientation = "landscape"
        ws_i.page_setup.fitToWidth = 1
        ws_i.page_setup.fitToHeight = 0
        ws_i.sheet_properties.pageSetUpPr.fitToPage = True

        total_row = iracp_table.iloc[-1]
        shortfall = -total_row["Excess/(Shortfall) of Ind AS 109 over IRACP_Cr"]
        note_row = ws_i.max_row + 2
        if shortfall > 0:
            ws_i.cell(row=note_row, column=1,
                      value=f"RESULT: Impairment Reserve required — Rs. {shortfall:.3f} Cr "
                            f"(Ind AS 109 provision is lower than IRACP floor by this amount).").font = Font(bold=True, color="FFC00000")
        else:
            ws_i.cell(row=note_row, column=1,
                      value="RESULT: Ind AS 109 provision meets or exceeds the IRACP floor. "
                            "No Impairment Reserve appropriation required.").font = Font(bold=True, color="FF006100")

        # ---- Assumptions register (security data availability) ----
        if iracp_assumptions is not None:
            write_dataframe_sheet(
                wb,
                sheet_name=f"IRACP_Assumptions_{iracp_period}",
                title=f"IRACP Computation — Assumptions Register — {iracp_period}",
                subtitle=("Where realisable security value was not supplied, doubtful assets are treated as "
                          "fully unsecured. Per para 32(2)(i) this can only OVERSTATE the prudential floor, "
                          "never understate it. Supply Secured_Value_Cr to refine."),
                df=iracp_assumptions,
            )

    wb.save(out_path)


# --------------------------------------------------------------------------
# MAIN
# --------------------------------------------------------------------------
def resolve_comparison_pairs(periods, compare_args, all_pairs: bool):
    """
    Decides WHICH period pairs to analyse.

    periods      : periods in the order they appear in the input file
    compare_args : list of "FROM:TO" strings from --compare (may be empty)
    all_pairs    : if True, every ordered combination of periods

    Precedence:
      1. --compare FROM:TO [FROM:TO ...]   -> exactly those pairs, in order
      2. --all-pairs                       -> every ordered pair (P1 earlier
                                              than P2 by file order)
      3. default                           -> consecutive pairs only

    The pairs need NOT be consecutive: Q3 vs Q1, Q4 vs Q2, or FY-end vs
    FY-end are all valid. Order within a pair is normalised so the earlier
    period (by position in the file) is always the "from" period, which
    keeps the migration matrix and the Ind AS 107 reconciliation
    directionally correct regardless of how the pair was typed.
    """
    index_of = {p: i for i, p in enumerate(periods)}

    if compare_args:
        pairs = []
        for spec in compare_args:
            if ":" not in spec:
                raise ValueError(
                    f"--compare expects FROM:TO (e.g. Q1FY26:Q3FY26), got '{spec}'"
                )
            a, b = [s.strip() for s in spec.split(":", 1)]
            for p in (a, b):
                if p not in index_of:
                    raise ValueError(
                        f"Period '{p}' not found in the input file. "
                        f"Available periods: {', '.join(periods)}"
                    )
            if a == b:
                raise ValueError(f"--compare needs two different periods, got '{spec}'")
            # normalise so the earlier period (by file order) is 'from'
            pairs.append((a, b) if index_of[a] < index_of[b] else (b, a))
        return pairs

    if all_pairs:
        return [(periods[i], periods[j])
                for i in range(len(periods))
                for j in range(i + 1, len(periods))]

    return list(zip(periods[:-1], periods[1:]))


# --------------------------------------------------------------------------
# PERIOD-END DATA REPOSITORY
# --------------------------------------------------------------------------
# The tool works off a folder in which each period-end extract is stored as
# its own file. It scans that folder, works out which period each file holds,
# and lets the user pick any two of them to compare - current year, prior
# year, or across years. Nothing about the periods is hard-coded.
#
#   data/
#     Q1FY26.csv        <- one file per period end
#     Q2FY26.csv
#     Q3FY26.xlsx
#     Q4FY25.xlsb
#
# The period label is taken from the Period column inside the file. If that
# column is absent, the file name (without extension) is used as the period
# label, so a bare "Q3FY26.csv" with no Period column still works.
# A single file containing several periods is also supported - every period
# inside it is registered.

SUPPORTED_EXTS = (".csv", ".xlsx", ".xlsm", ".xls", ".xlsb")


def scan_period_repository(repo_dir: str):
    """
    Scans a folder of period-end files.
    Returns (periods_index, frames) where:
      periods_index : dict {period_label -> source file path}
      frames        : dict {period_label -> DataFrame for that period}
    Periods are sorted using sort_period_key so FY/quarter labels order
    chronologically rather than alphabetically.
    """
    repo_dir = os.path.abspath(repo_dir)

    if not os.path.isdir(repo_dir):
        # Create it rather than failing, and tell the user exactly what to do.
        try:
            os.makedirs(repo_dir, exist_ok=True)
            created = True
        except OSError as e:
            raise ValueError(
                f"Could not find or create the data folder:\n    {repo_dir}\n({e})"
            )
        raise ValueError(
            "No period-end data folder was found, so one has just been created for you:\n"
            f"    {repo_dir}\n\n"
            "Next step: put ONE file per period end into that folder, for example\n"
            "    Q1FY26.csv   Q2FY26.xlsx   Q3FY26.csv   Q4FY26.xlsb\n\n"
            "Each file needs these columns:\n"
            "    LoanID, CustomerName, Stage, Outstanding_Cr, Provision_Cr\n"
            "    (optionally: Period, DPD, Secured_Value_Cr, Loss_Flag)\n\n"
            "If a file has no Period column, its file name is used as the period label.\n"
            "Then run this script again."
        )

    files = sorted(
        f for f in os.listdir(repo_dir)
        if os.path.splitext(f)[1].lower() in SUPPORTED_EXTS
        and not f.startswith("~$")
    )
    if not files:
        raise ValueError(
            f"The data folder exists but is empty:\n    {repo_dir}\n\n"
            "Put one file per period end into it (e.g. Q1FY26.csv, Q2FY26.xlsx),\n"
            f"then run again. Supported types: {', '.join(SUPPORTED_EXTS)}"
        )

    periods_index, frames, writeoff_files = {}, {}, []
    for fname in files:
        fpath = os.path.join(repo_dir, fname)
        if is_writeoff_file(fname):
            writeoff_files.append(fpath)
            continue
        try:
            fdf = load_data(fpath, require_period=False)
        except Exception as e:
            print(f"  ! Skipping '{fname}': {e}")
            continue

        if "Period" in fdf.columns:
            fdf["Period"] = fdf["Period"].astype(str).str.strip()
            file_periods = list(dict.fromkeys(fdf["Period"]))
        else:
            label = os.path.splitext(fname)[0].strip()
            fdf = fdf.copy()
            fdf["Period"] = label
            file_periods = [label]

        for per in file_periods:
            if per in periods_index:
                print(f"  ! Period '{per}' already loaded from "
                      f"'{os.path.basename(periods_index[per])}'; ignoring copy in '{fname}'.")
                continue
            periods_index[per] = fpath
            frames[per] = fdf[fdf["Period"] == per].copy()

    if not periods_index:
        raise ValueError(f"No usable period data found in '{repo_dir}'.")

    ordered = sorted(periods_index.keys(), key=sort_period_key)

    # ---- write-off files ----
    writeoff_frames = {}
    for wpath in writeoff_files:
        try:
            wdf = load_data(wpath, require_period=False, minimal=True)
        except Exception as e:
            print(f"  ! Skipping write-off file '{os.path.basename(wpath)}': {e}")
            continue
        if "LoanID" not in wdf.columns:
            print(f"  ! Skipping write-off file '{os.path.basename(wpath)}': no LoanID column")
            continue
        if "Period" in wdf.columns:
            wdf["Period"] = wdf["Period"].astype(str).str.strip()
            groups = [(per, g) for per, g in wdf.groupby("Period")]
        else:
            per = extract_period_from_name(wpath, ordered)
            if per is None:
                print(f"  ! Skipping write-off file '{os.path.basename(wpath)}': "
                      f"no Period column and the file name does not name a known period")
                continue
            groups = [(per, wdf)]
        for per, g in groups:
            if per not in ordered:
                print(f"  ! Write-off file names period '{per}', which has no data file. Ignored.")
                continue
            writeoff_frames[per] = pd.concat([writeoff_frames[per], g]) if per in writeoff_frames else g.copy()
            print(f"  + Write-off file loaded for {per}: {len(g)} loan(s) "
                  f"from '{os.path.basename(wpath)}'")

    return ({p: periods_index[p] for p in ordered},
            {p: frames[p] for p in ordered},
            writeoff_frames)


def sort_period_key(label: str):
    """
    Orders period labels chronologically rather than alphabetically.
    Understands the common Indian FY conventions:
        Q1FY26, Q2FY26 ... ; FY26 / FY2026 ; Mar-26, 31-Mar-2026 ;
        H1FY26 ; 2026-03-31
    Anything unrecognised sorts last, alphabetically, so the tool never
    crashes on an unexpected label - it just orders it at the end.
    """
    s = str(label).strip().upper().replace(" ", "")

    m = re.match(r"^Q([1-4])FY(\d{2,4})$", s)          # Q3FY26 / Q3FY2026
    if m:
        q, y = int(m.group(1)), int(m.group(2))
        y = 2000 + y if y < 100 else y
        return (0, y, q)

    m = re.match(r"^H([12])FY(\d{2,4})$", s)           # H1FY26
    if m:
        h, y = int(m.group(1)), int(m.group(2))
        y = 2000 + y if y < 100 else y
        return (0, y, h * 2)

    m = re.match(r"^FY(\d{2,4})$", s)                  # FY26 (full year)
    if m:
        y = int(m.group(1))
        y = 2000 + y if y < 100 else y
        return (0, y, 4)

    m = re.match(r"^(\d{4})-(\d{2})-(\d{2})$", s)     # 2026-03-31
    if m:
        return (0, int(m.group(1)), int(m.group(2)) / 12)

    months = {"JAN": 1, "FEB": 2, "MAR": 3, "APR": 4, "MAY": 5, "JUN": 6,
              "JUL": 7, "AUG": 8, "SEP": 9, "OCT": 10, "NOV": 11, "DEC": 12}
    m = re.match(r"^(?:\d{1,2}-)?([A-Z]{3})-?(\d{2,4})$", s)   # Mar-26 / 31-Mar-2026
    if m and m.group(1) in months:
        y = int(m.group(2))
        y = 2000 + y if y < 100 else y
        return (0, y, months[m.group(1)] / 12)

    return (1, 0, 0, s)


def prompt_period_selection(periods_index, frames):
    """
    Shows the periods available in the repository and asks the user which two
    to compare. Returns (period_from, period_to) with the earlier period
    first. Accepts either the list numbers ("1 3") or the labels
    ("Q1FY26 Q3FY26").
    """
    ordered = list(periods_index.keys())

    print("\nPeriod-end data available:")
    print(f"  {'#':<4}{'Period':<14}{'Loans':>8}{'Outstanding (Rs. Cr)':>24}   Source file")
    print("  " + "-" * 82)
    for i, per in enumerate(ordered, 1):
        fdf = frames[per]
        print(f"  {i:<4}{per:<14}{fdf['LoanID'].nunique():>8}"
              f"{fdf['Outstanding_Cr'].sum():>24,.2f}   {os.path.basename(periods_index[per])}")
    print()

    if len(ordered) < 2:
        raise ValueError("At least two periods are needed to run a comparison.")

    def _resolve(token):
        token = token.strip()
        if token.isdigit() and 1 <= int(token) <= len(ordered):
            return ordered[int(token) - 1]
        for per in ordered:
            if per.upper() == token.upper():
                return per
        return None

    while True:
        raw = input("Which two periods do you want to compare? "
                    "(enter two numbers or labels, e.g. '1 3' or 'Q1FY26 Q3FY26'): ").strip()
        if not raw:
            print("  Please enter two periods.\n")
            continue
        tokens = raw.replace(",", " ").replace(":", " ").split()
        if len(tokens) != 2:
            print("  Please enter exactly two periods.\n")
            continue
        a, b = _resolve(tokens[0]), _resolve(tokens[1])
        if a is None or b is None:
            bad = tokens[0] if a is None else tokens[1]
            print(f"  '{bad}' is not one of the periods listed above.\n")
            continue
        if a == b:
            print("  Please choose two different periods.\n")
            continue
        # normalise: earlier period first
        i_a, i_b = ordered.index(a), ordered.index(b)
        return (a, b) if i_a < i_b else (b, a)


def main():
    parser = argparse.ArgumentParser(
        description="ECL Stage Migration Matrix & Commentary Narrator (NBFC-Middle Layer)",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""Examples:
  # default - consecutive periods (Q1 vs Q2, Q2 vs Q3, ...)
  python ecl_migration_matrix.py data.csv

  # a specific non-consecutive comparison: Q1 vs Q3
  python ecl_migration_matrix.py data.csv --compare Q1FY26:Q3FY26

  # several comparisons in one run
  python ecl_migration_matrix.py data.csv --compare Q1FY26:Q3FY26 Q2FY26:Q4FY26

  # every possible period pair
  python ecl_migration_matrix.py data.csv --all-pairs

  # run the IRACP disclosure as at a period other than the latest
  python ecl_migration_matrix.py data.csv --compare Q1FY26:Q3FY26 --iracp-period Q3FY26
""")
    parser.add_argument("--data-dir", metavar="DIR", default=None,
                        help="Folder holding one file per period end (default: ./data). "
                             "The tool scans it, lists the periods, and asks which to compare.")
    parser.add_argument("--input-file", metavar="FILE", default=None,
                        help="Alternative to --data-dir: a single file containing several periods.")
    parser.add_argument("--ai", action="store_true", help="Also attempt AI-enhanced narrative (needs internet + ANTHROPIC_API_KEY)")
    parser.add_argument("--compare", nargs="+", metavar="FROM:TO", default=None,
                        help="One or more period pairs to compare, e.g. Q1FY26:Q3FY26. "
                             "Pairs need NOT be consecutive. Overrides the default.")
    parser.add_argument("--all-pairs", action="store_true",
                        help="Analyse every possible ordered pair of periods in the file.")
    parser.add_argument("--iracp-period", metavar="PERIOD", default=None,
                        help="Period as at which to run the RBI IRACP disclosure. "
                             "Default: the latest period in the comparisons being run.")
    parser.add_argument("--output-dir", metavar="DIR", default=None,
                        help="Base folder for the input/ archive and output/ run folders. "
                             "Default: the folder containing the input file.")
    parser.add_argument("--list-periods", action="store_true",
                        help="List the periods found in the input file and exit.")
    args = parser.parse_args()

    # ------------------------------------------------------------------
    # LOAD: either from a period-end repository folder (default mode) or
    # from one multi-period file (--input-file, for backward compatibility)
    # ------------------------------------------------------------------
    if args.input_file:
        source_desc = os.path.abspath(args.input_file)
        df = load_data(args.input_file)
        periods = sorted(dict.fromkeys(df["Period"]), key=sort_period_key)
        periods_index = {p: args.input_file for p in periods}
        frames = {p: df[df["Period"] == p].copy() for p in periods}
        writeoff_frames = {}
    else:
        # Default the data folder to one sitting NEXT TO THIS SCRIPT/EXE, not
        # to the current working directory. IDLE, Task Scheduler and
        # double-click launches often run with a different working
        # directory, which would otherwise send the tool looking in the
        # wrong place. SCRIPT_DIR itself already accounts for the
        # PyInstaller-frozen case (see its definition near the top of this
        # file), so this line needs no separate frozen-check of its own.
        repo_dir = args.data_dir or os.path.join(SCRIPT_DIR, "data")
        source_desc = os.path.abspath(repo_dir)
        print(f"Scanning period-end data folder: {source_desc}")
        periods_index, frames, writeoff_frames = scan_period_repository(repo_dir)
        periods = list(periods_index.keys())
        df = pd.concat([frames[p] for p in periods], ignore_index=True)

    if args.list_periods:
        print(f"\nPeriods available in {source_desc}:")
        print(f"  {'#':<4}{'Period':<14}{'Loans':>8}{'Outstanding (Rs. Cr)':>24}   Source file")
        print("  " + "-" * 82)
        for i, p in enumerate(periods, 1):
            fdf = frames[p]
            print(f"  {i:<4}{p:<14}{fdf['LoanID'].nunique():>8}"
                  f"{fdf['Outstanding_Cr'].sum():>24,.2f}   {os.path.basename(periods_index[p])}")
        return

    if len(periods) < 2:
        raise ValueError(
            f"Only one period ({periods[0] if periods else 'none'}) is available. "
            "At least two period-end files are needed to run a comparison."
        )

    # ------------------------------------------------------------------
    # SELECT PERIODS: interactive by default; --compare / --all-pairs to
    # skip the prompt (useful for scheduled or batch runs)
    # ------------------------------------------------------------------
    if args.compare or args.all_pairs:
        pairs = resolve_comparison_pairs(periods, args.compare, args.all_pairs)
    else:
        pairs = [prompt_period_selection(periods_index, frames)]

    result = run_pipeline(
        periods_index=periods_index, frames=frames, writeoff_frames=writeoff_frames,
        periods=periods, pairs=pairs, source_desc=source_desc,
        iracp_period=args.iracp_period, use_ai=args.ai,
        output_dir=args.output_dir, input_file=args.input_file, data_dir=args.data_dir,
    )
    return result


# --------------------------------------------------------------------------
# SHARED ENGINE ENTRY POINT
# --------------------------------------------------------------------------
# Everything from period selection onward - write-off handling, migration
# matrices, statutory disclosures, commentary, and writing the workbook/
# commentary/run-log to disk - lives in this ONE function. Both the
# command-line main() above and the optional Tkinter GUI (ecl_gui.py) call
# THIS function and nothing else, so the GUI can never compute a different
# answer than the command-line tool: it can only choose which data folder
# and which period pairs to run, never how the numbers are derived.
def run_pipeline(periods_index: dict, frames: dict, writeoff_frames: dict, periods: list,
                  pairs: list, source_desc: str, iracp_period: str = None,
                  use_ai: bool = False, output_dir: str = None, input_file: str = None,
                  data_dir: str = None) -> str:
    """
    Runs the full analysis for an already-resolved set of period pairs and
    writes the output workbook, commentary text file and run log.

    Parameters mirror what main() already has in hand after scanning/loading
    and resolving --compare/--all-pairs/the interactive prompt - this
    function does no scanning or period-pair resolution of its own.

    Returns the path to the timestamped output run folder.
    """
    safe_print("\nComparison(s) being analysed: "
               + "; ".join(f"{a} {ARROW} {b}" for a, b in pairs))

    df = pd.concat([frames[p] for p in periods], ignore_index=True)

    # Restrict everything downstream to ONLY the periods actually selected,
    # so the workbook contains no sheets for periods the user did not ask for.
    selected_periods = [p for p in periods if any(p in pair for pair in pairs)]
    df = df[df["Period"].isin(selected_periods)].copy()
    print(f"Periods included in this report: {', '.join(selected_periods)}\n")

    # ------------------------------------------------------------------
    # WRITE-OFFS: identify them from the RAW data first (so the pre-write-off
    # stage and provision are captured for the disclosure), THEN remove those
    # loans from the ACTIVE portfolio from their write-off period onward.
    #
    # Without this step a written-off loan would be counted BOTH in the
    # period's closing stage totals (it is still sitting in the raw file,
    # merely flagged) AND as a deduction in the write-off line - double
    # counting the same rupee. Once removed from the active portfolio it
    # behaves exactly like a loan that exited the book: it drops out of
    # every period from the write-off period onward, feeding the "closed"
    # side of the migration matrix, which the Ind AS 107 reconciliation
    # then correctly attributes to "Amounts written off" rather than
    # ordinary repayment.
    # ------------------------------------------------------------------
    period_rank = {p: i for i, p in enumerate(selected_periods)}
    has_inline_flag = "WriteOff_Flag" in df.columns
    has_writeoff_files = bool(writeoff_frames)
    print("Write-off detection:")
    print(f"  - Inline 'WriteOff_Flag' column in period files: "
          f"{'found' if has_inline_flag else 'NOT found'}")
    print(f"  - Separate write-off file(s) in the data folder: "
          f"{'found for ' + ', '.join(sorted(writeoff_frames.keys())) if has_writeoff_files else 'NONE found'}"
          + ("" if has_writeoff_files else
             f" (looked for file names containing: {', '.join(WRITEOFF_FILE_MARKERS)})"))
    if not has_inline_flag and not has_writeoff_files:
        print("  -> No write-off data available by either route. The write-off section of the "
              "commentary, the Ind AS 107 'Amounts written off' line, and the migration sheet's "
              "write-off entry will all be zero/blank for this run - this is expected, not an error, "
              "unless you intended to supply write-off data. See the script header for the exact "
              "column name and file-naming convention.")

    writeoffs_by_period = {per: collect_writeoffs(df, writeoff_frames, per) for per in selected_periods}
    for per in selected_periods:
        n = len(writeoffs_by_period[per])
        if n:
            amt = writeoffs_by_period[per]["WriteOff_Amount_Cr"].sum()
            print(f"  - {per}: {n} loan(s) identified as written off, Rs. {amt:,.3f} Cr")
    print()

    first_wo_period_for_loan = {}
    for per in selected_periods:            # selected_periods is chronological
        for lid in writeoffs_by_period[per]["LoanID"]:
            first_wo_period_for_loan.setdefault(lid, per)

    if first_wo_period_for_loan:
        wo_rank = df["LoanID"].map(first_wo_period_for_loan).map(period_rank)
        period_rank_col = df["Period"].map(period_rank)
        exclude_mask = wo_rank.notna() & (period_rank_col >= wo_rank)
        df_active = df[~exclude_mask].copy()
        print(f"Write-offs applied: {len(first_wo_period_for_loan)} loan(s) removed from the "
              f"active portfolio from their write-off period onward.\n")
    else:
        df_active = df

    print("Product-wise and Branch-wise sheets:")
    print(f"  - 'Product' column: {'found -> Product summary sheet will be generated' if 'Product' in df_active.columns else 'NOT found in the input data -> Product summary sheet will be SKIPPED'}")
    print(f"  - 'Branch' column: {'found -> Branch filter sheet will be generated' if 'Branch' in df_active.columns else 'NOT found in the input data -> Branch filter sheet will be SKIPPED'}")
    print()

    summary = compute_period_summary(df_active)
    summary["_o"] = summary["Period"].map({p: i for i, p in enumerate(selected_periods)})
    summary = summary.sort_values("_o").drop(columns="_o").reset_index(drop=True)

    transitions = []
    migrations_for_commentary = []
    indas107_tables = []
    dpd_migrations = {}
    for p_from, p_to in pairs:
        mcount, mamount, mprov, detail, closed, new = build_migration_matrix(df_active, p_from, p_to)
        transitions.append((p_from, p_to, mcount, mamount, mprov, detail, closed, new))
        migrations_for_commentary.append((mcount, mamount, detail, closed, new))
        wo_to = writeoffs_by_period.get(p_to)
        recon_df = compute_indas107_reconciliation(df_active, p_from, p_to, writeoffs=wo_to)
        indas107_tables.append((p_from, p_to, recon_df))
        dpd_mig = build_dpd_migration(df_active, p_from, p_to)
        if dpd_mig is not None:
            dpd_migrations[(p_from, p_to)] = dpd_mig

    iracp_table = iracp_method_note = None
    iracp_assumptions = iracp_loan_detail = None
    resolved_iracp_period = None
    if "DPD" in df.columns:
        if iracp_period:
            if iracp_period not in selected_periods:
                raise ValueError(
                    f"IRACP period '{iracp_period}' is not among the periods selected "
                    f"for this run ({', '.join(selected_periods)})."
                )
            resolved_iracp_period = iracp_period
        else:
            # latest period actually involved in the selected comparisons
            resolved_iracp_period = max((p for pair in pairs for p in pair),
                                        key=lambda p: selected_periods.index(p))
        iracp_table, iracp_method_note, iracp_assumptions, iracp_loan_detail = \
            compute_iracp_comparison(df_active, resolved_iracp_period)

    latest_wo = writeoffs_by_period.get(pairs[-1][1]) if writeoffs_by_period else None
    latest_product_summary = None
    if "Product" in df_active.columns:
        latest_product_summary = build_group_summary(df_active, pairs[-1][0], pairs[-1][1], "Product")
    rule_based = generate_rule_based_commentary(summary, migrations_for_commentary, pairs,
                                                 writeoffs=latest_wo, product_summary=latest_product_summary)
    print(rule_based)

    final_commentary = rule_based
    if use_ai:
        print("\n\nAttempting AI-enhanced narrative...\n")
        ai_version = generate_ai_commentary(summary, rule_based)
        print(ai_version)
        final_commentary = rule_based + "\n\n\n" + ai_version

    # ------------------------------------------------------------------
    # FILE MANAGEMENT: archive the input, timestamp the output
    # ------------------------------------------------------------------
    #   <base_dir>/input/   - a dated copy of every input file processed,
    #                         so the exact data behind any report is retained
    #   <base_dir>/output/  - one timestamped sub-folder per run, holding the
    #                         workbook, the commentary text and a run log
    # This gives a clean audit trail: any output folder can be traced back to
    # the precise input file that produced it.
    if output_dir:
        base_dir = os.path.abspath(output_dir)
    elif input_file:
        base_dir = os.path.dirname(os.path.abspath(input_file)) or "."
    else:
        base_dir = os.path.dirname(os.path.abspath(
            data_dir or os.path.join(SCRIPT_DIR, "data"))) or SCRIPT_DIR

    input_archive_dir = os.path.join(base_dir, "input_archive")
    run_stamp = datetime.now().strftime("%Y-%m-%d_%H%M%S")
    pair_tag = "_".join(f"{a}-vs-{b}" for a, b in pairs)[:60]
    output_run_dir = os.path.join(base_dir, "output", f"{pair_tag}_{run_stamp}")
    os.makedirs(input_archive_dir, exist_ok=True)
    os.makedirs(output_run_dir, exist_ok=True)

    # Archive a dated copy of ONLY the source files behind the selected
    # periods (never modify the originals), and hash each for the run log.
    source_files, archived_inputs = [], []
    for per in selected_periods:
        sf = periods_index[per]
        if sf not in source_files:
            source_files.append(sf)
    for sf in source_files:
        stem = os.path.splitext(os.path.basename(sf))[0]
        ext = os.path.splitext(sf)[1]
        dest = os.path.join(input_archive_dir, f"{stem}_{run_stamp}{ext}")
        shutil.copy2(sf, dest)
        archived_inputs.append(dest)

    base = pair_tag or "ECL"
    out_path = os.path.join(output_run_dir, f"{base}_ECL_analysis_{run_stamp}.xlsx")
    txt_path = os.path.join(output_run_dir, f"{base}_commentary_{run_stamp}.txt")
    log_path = os.path.join(output_run_dir, f"run_log_{run_stamp}.txt")

    write_excel(summary, final_commentary, transitions, out_path,
                indas107_tables=indas107_tables, iracp_table=iracp_table, iracp_period=resolved_iracp_period,
                iracp_method_note=iracp_method_note, iracp_assumptions=iracp_assumptions,
                iracp_loan_detail=iracp_loan_detail, dpd_migrations=dpd_migrations,
                writeoffs_by_period=writeoffs_by_period, df_active=df_active)

    # encoding="utf-8" is required here: commentary/log text can contain the
    # arrow character (\u2192) used in "Q1FY26 -> Q2FY26" style labels, and
    # opening a file without an explicit encoding on Windows falls back to
    # the system's default codepage (commonly cp1252), which cannot encode
    # that character and raises a UnicodeEncodeError partway through the
    # write - this bit both files explicitly to UTF-8 regardless of the
    # machine's locale.
    with open(txt_path, "w", encoding="utf-8") as f:
        f.write(final_commentary)

    with open(log_path, "w", encoding="utf-8") as f:
        f.write("ECL Migration & Disclosure Tool — Run Log\n")
        f.write("=" * 60 + "\n")
        f.write(f"Run timestamp        : {datetime.now().strftime('%d-%b-%Y %H:%M:%S')}\n")
        f.write(f"Data source          : {source_desc}\n")
        f.write("Source files used    :\n")
        for sf, ai in zip(source_files, archived_inputs):
            f.write(f"    {os.path.basename(sf)}\n")
            f.write(f"      SHA-256  : {file_sha256(sf)}\n")
            f.write(f"      archived : {ai}\n")
        f.write(f"Periods available    : {', '.join(periods)}\n")
        f.write(f"Periods SELECTED     : {', '.join(selected_periods)}\n")
        f.write(f"Comparisons analysed : {'; '.join(f'{a} -> {b}' for a, b in pairs)}\n")
        f.write(f"IRACP period         : {resolved_iracp_period or 'not run (no DPD column)'}\n")
        f.write(f"IRACP basis          : NBFC-Middle Layer, RBI IRACP Directions, 2025\n")
        f.write(f"DPD buckets          : {', '.join(DPD_BUCKET_NAMES)}\n")
        f.write(f"AI narrative         : {'yes' if use_ai else 'no (deterministic only)'}\n")
        f.write(f"Workbook written     : {out_path}\n")

    print(f"\n{'=' * 66}")
    for ai in archived_inputs:
        print(f"Input archived    : {ai}")
    print(f"Output folder     : {output_run_dir}")
    print(f"  - {os.path.basename(out_path)}")
    print(f"  - {os.path.basename(txt_path)}")
    print(f"  - {os.path.basename(log_path)}")
    print(f"{'=' * 66}")

    return output_run_dir


if __name__ == "__main__":
    # Friendly top-level error handling: setup problems (missing folder,
    # missing columns, bad period selection) are shown as a clear message
    # rather than a Python traceback. Unexpected faults still raise in full
    # so they can be diagnosed.
    try:
        main()
    except ValueError as e:
        print("\n" + "=" * 70)
        print("COULD NOT COMPLETE THE RUN")
        print("=" * 70)
        print(e)
        print("=" * 70)
        sys.exit(1)
    except KeyboardInterrupt:
        print("\nCancelled.")
        sys.exit(1)
